"""ROScribe — Scholar's Archive workspace (NiceGUI · Concept 2).

Panes:
  Left   — PDF       : the judgment document (inline viewer)
  Center — Breakdown : reliable metadata (parties, bench) + LLM analysis;
                       topic chips are clickable -> related judgements
  Right  — Library   : live keyword search, By-Justice + By-area filters,
                       and judgements grouped by Year (default).
"""

from __future__ import annotations

import json
import os
import re
import secrets
import sqlite3
import sys
import time
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from src.config import REPO_ROOT, settings  # noqa: E402
from src.ingest import extract_bench, merge_benches  # noqa: E402
from src.schema import NOT_AVAILABLE  # noqa: E402
from src.store import (  # noqa: E402
    LEGAL_AREAS, area_search, citation_search_terms, combined_search, embedder_ready,
    keyword_search, resolve_citation,
)

from fastapi import Request  # noqa: E402
from fastapi.responses import FileResponse, RedirectResponse, Response  # noqa: E402
from starlette.middleware.base import BaseHTTPMiddleware  # noqa: E402
from nicegui import Client, app, run, ui  # noqa: E402

JUDGE_DIR = REPO_ROOT / "data" / "sc_judgements"


def _safe_child(base: Path, name: str) -> Path | None:
    """Resolve `name` strictly inside `base`, or None if it escapes (path
    traversal). Guards the file-serving routes against `../`, absolute paths,
    symlinks, and percent-encoded variants that Starlette has already decoded
    into the path parameter."""
    if not name or "\x00" in name:
        return None
    try:
        base = base.resolve()
        candidate = (base / name).resolve()
        candidate.relative_to(base)          # raises ValueError if outside base
    except (ValueError, OSError):
        return None
    return candidate


@app.get("/pdf/{name}")
def serve_pdf(name: str):
    p = _safe_child(JUDGE_DIR, name)
    if p is None or not p.is_file() or p.suffix.lower() != ".pdf":
        return Response(status_code=404)
    safe_dl = re.sub(r"[^A-Za-z0-9._-]+", "_", p.name)
    return FileResponse(str(p), media_type="application/pdf",
                        headers={"Content-Disposition": f'inline; filename="{safe_dl}"'})


@app.get("/logo/{name}")
def serve_logo(name: str):
    p = _safe_child(REPO_ROOT / "data" / "logos", name)
    if p is None or not p.is_file():
        return Response(status_code=404)
    return FileResponse(str(p))


# -------------------- access control (closed user base) ------------------ #
def _parse_users(raw: str) -> dict[str, str]:
    out: dict[str, str] = {}
    for pair in raw.split(","):
        if ":" in pair:
            u, p = pair.split(":", 1)
            out[u.strip()] = p.strip()
    return out


USERS = _parse_users(os.getenv("ROSCRIBE_USERS", ""))
_DEFAULT_SECRET = "roscribe-change-this-secret"
STORAGE_SECRET = os.getenv("ROSCRIBE_STORAGE_SECRET", _DEFAULT_SECRET)
if STORAGE_SECRET == _DEFAULT_SECRET:
    # A known secret lets anyone forge a signed session cookie → full auth bypass.
    # scripts/roscribe.sh writes a random one into .env on first start; warn if not.
    print("⚠️  ROSCRIBE_STORAGE_SECRET is unset — using the INSECURE default. "
          "Set it in .env before exposing the app (run via scripts/roscribe.sh).")
UNRESTRICTED = {"/login", "/demo"}

# Brute-force throttle: per-username failed-attempt timestamps (in-memory). After
# _LOGIN_MAX failures within _LOGIN_WINDOW seconds, further attempts on that
# username are delayed/blocked until the window slides. Survives cookie-clearing
# (keyed server-side, not in the session) and can't permanently lock a user out.
_LOGIN_FAILS: dict[str, list[float]] = {}
_LOGIN_WINDOW = 300.0
_LOGIN_MAX = 6


def _login_recent_fails(username: str) -> int:
    now = time.monotonic()
    fails = [t for t in _LOGIN_FAILS.get(username, []) if now - t < _LOGIN_WINDOW]
    _LOGIN_FAILS[username] = fails
    return len(fails)


def _record_login_fail(username: str) -> None:
    _LOGIN_FAILS.setdefault(username, []).append(time.monotonic())


class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        if not app.storage.user.get("authenticated", False):
            if request.url.path in Client.page_routes.values() and request.url.path not in UNRESTRICTED:
                app.storage.user["referrer_path"] = request.url.path
                return RedirectResponse("/login")
        return await call_next(request)


app.add_middleware(AuthMiddleware)


@ui.page("/login", title="ROS", favicon="data/logos/logo_emblem.png")
def login():
    if app.storage.user.get("authenticated", False):
        return RedirectResponse("/")

    def attempt():
        user = (username.value or "").strip()
        if _login_recent_fails(user) >= _LOGIN_MAX:
            ui.notify("Too many attempts — wait a few minutes and try again.", color="negative")
            return
        # Timing-safe compare so response time doesn't leak whether the username
        # exists or how much of the password matched.
        expected = USERS.get(user, "")
        ok = bool(password.value) and bool(expected) and \
            secrets.compare_digest(str(password.value), str(expected))
        if ok:
            _LOGIN_FAILS.pop(user, None)
            app.storage.user.update({"username": user, "authenticated": True})
            ui.navigate.to(app.storage.user.get("referrer_path", "/"))
        else:
            _record_login_fail(user)
            ui.notify("Invalid credentials", color="negative")

    with ui.card().classes("absolute-center w-80 items-stretch"):
        ui.label("⚖️ ROScribe").classes("text-xl font-bold self-center")
        ui.label("Supreme Court of Sri Lanka — legal research").classes("text-xs text-gray-500 self-center mb-2")
        username = ui.input("Username").props("outlined dense").on("keydown.enter", attempt)
        password = ui.input("Password", password=True, password_toggle_button=True).props("outlined dense").on("keydown.enter", attempt)
        ui.button("Log in", on_click=attempt).props("color=primary")
        ui.button("Try the demo →", on_click=lambda: ui.navigate.to("/demo")).props("flat dense").classes("self-center")
    return None


# ------------------------------ data ------------------------------------- #
def _con() -> sqlite3.Connection:
    return sqlite3.connect(settings.sqlite_path)


def init_db():
    con = _con()
    con.execute("""
        CREATE TABLE IF NOT EXISTS bookmarks (
            username TEXT,
            case_no TEXT,
            PRIMARY KEY (username, case_no)
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS annotations (
            username TEXT,
            case_no TEXT,
            notes TEXT,
            PRIMARY KEY (username, case_no)
        )
    """)
    con.commit()
    con.close()

init_db()


def is_bookmarked(user: str, case_no: str) -> bool:
    con = _con()
    row = con.execute("SELECT 1 FROM bookmarks WHERE username=? AND case_no=?", (user, case_no)).fetchone()
    con.close()
    return row is not None


def toggle_bookmark(user: str, case_no: str) -> bool:
    con = _con()
    if is_bookmarked(user, case_no):
        con.execute("DELETE FROM bookmarks WHERE username=? AND case_no=?", (user, case_no))
        status = False
    else:
        con.execute("INSERT OR REPLACE INTO bookmarks (username, case_no) VALUES (?, ?)", (user, case_no))
        status = True
    con.commit()
    con.close()
    return status


def get_bookmarks(user: str) -> list[str]:
    con = _con()
    rows = con.execute("SELECT case_no FROM bookmarks WHERE username=? ORDER BY case_no ASC", (user,)).fetchall()
    con.close()
    return [r[0] for r in rows]


def get_annotation(user: str, case_no: str) -> str:
    con = _con()
    row = con.execute("SELECT notes FROM annotations WHERE username=? AND case_no=?", (user, case_no)).fetchone()
    con.close()
    return row[0] if row else ""


def save_annotation(user: str, case_no: str, text: str):
    con = _con()
    con.execute("INSERT OR REPLACE INTO annotations (username, case_no, notes) VALUES (?, ?, ?)", (user, case_no, text))
    con.commit()
    con.close()


def get_citing_cases(case_no: str) -> list[str]:
    con = _con()
    rows = con.execute("SELECT case_no FROM analyses WHERE json LIKE ?", (f"%{case_no}%",)).fetchall()
    con.close()
    return [r[0] for r in rows if r[0] != case_no]


def _jl(s):
    try:
        return json.loads(s) if s else []
    except Exception:
        return []


def case_meta(case_no: str) -> dict:
    con = _con()
    row = con.execute(
        "SELECT case_no, date, parties, judges, keywords, legislation, filename "
        "FROM judgements WHERE case_no=? LIMIT 1", (case_no,)
    ).fetchone()
    con.close()
    if not row:
        return {}
    return {"case_no": row[0], "date": row[1] or "", "parties": row[2] or "",
            "judges": _jl(row[3]), "keywords": _jl(row[4]), "legislation": _jl(row[5]), "filename": row[6]}


_BENCH_CACHE: dict[str, list[str]] = {}


def bench_for(case_no: str, meta: dict) -> list[str]:
    """Full panel of judges for a case.

    The scrape only records the *authoring* judge, so we parse the real coram
    straight from the judgment text (`extract_bench`) and fall back to the
    scrape `judges` metadata when extraction finds nothing. Cached per case_no
    so each judgment is parsed at most once per process."""
    if case_no in _BENCH_CACHE:
        return _BENCH_CACHE[case_no]
    parsed: list[str] = []
    fn = meta.get("filename")
    if fn:
        pdf = JUDGE_DIR / fn
        if pdf.exists():
            try:
                parsed = extract_bench(str(pdf))
            except Exception:
                parsed = []
    # The scrape records only the *authoring* judge; the parsed coram (front
    # matter + signature block) records the concurring judges but sometimes
    # omits the author (who signs by role, not name). Merge both, surname-deduped,
    # so a 3- or 5-judge bench is complete even when each source alone is partial.
    meta_judges = [j for j in (meta.get("judges") or []) if str(j).strip()]
    bench = merge_benches(parsed, meta_judges) if parsed else meta_judges
    _BENCH_CACHE[case_no] = bench
    return bench


def get_breakdown(case_no: str):
    con = _con()
    row = con.execute("SELECT json FROM analyses WHERE case_no=?", (case_no,)).fetchone()
    con.close()
    return json.loads(row[0]) if row else None


def cases_by_judge(name: str) -> list[dict]:
    con = _con()
    rows = con.execute(
        "SELECT case_no, date, parties FROM judgements WHERE judges LIKE ? ORDER BY date DESC LIMIT 200",
        (f"%{name}%",)).fetchall()
    con.close()
    return [{"case_no": r[0], "date": r[1] or "", "snippet": (r[2] or "")[:100]} for r in rows]


def cases_by_keyword(area: str) -> list[dict]:
    con = _con()
    rows = con.execute(
        "SELECT case_no, date, parties FROM judgements WHERE keywords LIKE ? ORDER BY date DESC LIMIT 200",
        (f"%{area}%",)).fetchall()
    con.close()
    return [{"case_no": r[0], "date": r[1] or "", "snippet": (r[2] or "")[:100]} for r in rows]


_JUSTICES = None
_AREAS = None
_BY_YEAR = None


def distinct_justices() -> list[str]:
    """Deduped justice display names — one option per justice (variant spellings
    merged in src.store.justices_grouped)."""
    global _JUSTICES
    if _JUSTICES is None:
        from src.store import distinct_justices as _store_distinct_justices
        _JUSTICES = _store_distinct_justices()
    return _JUSTICES


def legal_areas() -> list[str]:
    return list(LEGAL_AREAS)  # curated real legal areas (see src/store.py)


def judgements_by_year() -> dict[str, list]:
    global _BY_YEAR
    if _BY_YEAR is None:
        con = _con()
        rows = con.execute("SELECT case_no, date FROM judgements").fetchall()
        con.close()
        by: dict[str, list] = {}
        for cn, date in rows:
            if date and date[:4].isdigit():
                y = date[:4]
            else:
                yrs = [int(x) for x in re.findall(r"(?:19|20)\d{2}", cn or "") if 1950 <= int(x) <= 2027]
                y = str(max(yrs)) if yrs else "Undated"
            by.setdefault(y, []).append((cn, date or ""))
        _BY_YEAR = by
    return _BY_YEAR


_BY_YEAR_MONTH = None
_MONTH_NAMES = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]


def judgements_by_year_month() -> dict[str, dict[str, list]]:
    global _BY_YEAR_MONTH
    if _BY_YEAR_MONTH is None:
        con = _con()
        rows = con.execute("SELECT case_no, date FROM judgements").fetchall()
        con.close()
        by: dict[str, dict[str, list]] = {}
        for cn, date in rows:
            if date and date[:4].isdigit():
                y = date[:4]
            else:
                yrs = [int(x) for x in re.findall(r"(?:19|20)\d{2}", cn or "") if 1950 <= int(x) <= 2027]
                y = str(max(yrs)) if yrs else "Undated"
            
            m_name = "Unknown Month"
            if date and len(date) >= 7 and date[5:7].isdigit():
                m_idx = int(date[5:7]) - 1
                if 0 <= m_idx < 12:
                    m_name = _MONTH_NAMES[m_idx]
            
            by.setdefault(y, {}).setdefault(m_name, []).append((cn, date or ""))
        _BY_YEAR_MONTH = by
    return _BY_YEAR_MONTH


def _available_years() -> list[str]:
    """Years present in the corpus (newest first) — options for the Year filter."""
    return sorted((y for y in judgements_by_year() if y != "Undated"), reverse=True)


def find_case(cited: str):
    """High-confidence citation -> corpus case_no (number-based; see
    src.store.resolve_citation). Party-name guessing is intentionally avoided."""
    return resolve_citation(cited)


def _safe_username() -> str:
    """Logged-in username, tolerant of a dropped/changed client session.

    A breakdown runs ~1-2 min via run.io_bound; if the websocket reconnects in
    that window the per-session user storage is gone and reading it raises
    'user storage ... should be created before accessing it'. Falling back to
    'anonymous' keeps the post-await re-render from crashing — which is what left
    the 'Analysing…' spinner stuck forever."""
    try:
        return app.storage.user.get("username", "anonymous")
    except Exception:
        return "anonymous"


def _web_search_url(query: str) -> str:
    """A Google search URL — the open-web fallback for any citation, Act, or
    constitutional article not in the local corpus, so it can still be looked up
    and cited from the original source."""
    from urllib.parse import quote_plus
    return "https://www.google.com/search?q=" + quote_plus((query or "").strip())


HEAD_CSS = """
<link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700&family=Lora:ital,wght@0,400;0,500;0,600;0,700;1,400&display=swap" rel="stylesheet">
<style>
  /* Theme: inspired by rechtspraak.nl / the Dutch Rijkshuisstijl —
     hemelblauw #01689b (primary), donkerblauw #154273 (headings), and a
     light governmental gray page. Flat, squared, generous white space. */
  body {
    background: #f3f3f3;
    font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
  }
  .case-title {
    font-family: 'Plus Jakarta Sans', sans-serif;
  }
  .pane-head {
    font-size: 0.75rem;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: #154273; /* donkerblauw */
    font-weight: 700;
    margin-bottom: 4px;
  }
  .doc-pane {
    background: #ffffff;
  }
  .sec {
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 0.85rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    color: #154273; /* donkerblauw heading */
    border-left: 3px solid #01689b; /* hemelblauw accent — the rechtspraak.nl signature */
    padding-left: 8px;
    margin-top: 1.2rem;
    margin-bottom: 0.4rem;
  }
  .body-text {
    font-family: 'Lora', Georgia, serif;
    font-size: 0.92rem;
    line-height: 1.6;
    color: #3c4043;
  }
  .chip {
    cursor: pointer;
    border-radius: 6px;
    font-family: 'Plus Jakarta Sans', sans-serif;
  }
  .q-item:hover {
    background: #f1f3f4;
  }
  /* Google style scrollbar */
  ::-webkit-scrollbar {
    width: 6px;
    height: 6px;
  }
  ::-webkit-scrollbar-track {
    background: transparent;
  }
  ::-webkit-scrollbar-thumb {
    background: #dadce0;
    border-radius: 10px;
  }
  ::-webkit-scrollbar-thumb:hover {
    background: #bdc1c6;
  }
  /* Dark Mode specific overrides */
  .body--dark {
    background: #121212 !important;
    color: #e8eaed !important;
  }
  .body--dark .sec {
    color: #8ab4f8 !important;
  }
  .body--dark .pane-head {
    color: #9aa0a6 !important;
  }
  .body--dark .body-text {
    color: #e8eaed !important;
  }
  .body--dark ::-webkit-scrollbar-thumb {
    background: #3c4043;
  }
  .body--dark ::-webkit-scrollbar-thumb:hover {
    background: #5f6368;
  }
  .body--dark .bg-white {
    background-color: #1e1e1e !important;
  }
  .body--dark .bg-gray-100 {
    background-color: #121212 !important;
  }
  .body--dark .bg-gray-50 {
    background-color: #2c2c2c !important;
  }
  .body--dark .border {
    border-color: #2c2c2c !important;
  }
  .body--dark .text-gray-900 {
    color: #f1f3f4 !important;
  }
  .body--dark .text-gray-800 {
    color: #e8eaed !important;
  }
  .body--dark .text-gray-700 {
    color: #dadce0 !important;
  }
  .body--dark .text-gray-600 {
    color: #bdc1c6 !important;
  }
  .body--dark .text-gray-500 {
    color: #9aa0a6 !important;
  }
  /* responsive: 3 panes side-by-side on desktop; one at a time + tab bar on mobile */
  .mobile-tabs {
    display: none;
  }
  .panes-row {
    height: calc(100vh - 56px - 30px);
  }
  @media (max-width: 900px) {
    .mobile-tabs {
      display: flex;
    }
    .panes-row {
      height: calc(100vh - 56px - 48px - 30px);
      padding: 8px !important;
      gap: 8px !important;
    }
    .pane {
      display: none !important;
      width: 100% !important;
    }
    .pane.active {
      display: flex !important;
    }
    .hide-narrow {
      display: none !important;
    }
  }
</style>
""";

_TREAT_COLOR = {"Distinguished": "orange", "Overruled": "red", "Applied": "green", "Followed": "green"}


def build_workspace(demo: bool = False):
    ui.add_head_html(HEAD_CSS)
    # Rechtspraak.nl / Rijkshuisstijl palette: hemelblauw primary, donkerblauw secondary.
    ui.colors(primary="#01689b", secondary="#154273")
    ui.dark_mode().disable()
    state = {"case": None, "page": None, "chat_open": False, "workspace_active": False}
    containers = {"bookmarks": None}

    def refresh_bookmarks_ui():
        if not containers["bookmarks"]:
            return
        containers["bookmarks"].clear()
        username = _safe_username()
        saved = get_bookmarks(username)
        if not saved:
            return
        with containers["bookmarks"]:
            exp = ui.expansion().classes("w-full mb-1").props("dense header-class='bg-blue-50 text-primary text-xs font-semibold rounded-md' expand-icon-class='text-primary'")
            with exp.add_slot('header'):
                with ui.row().classes("items-center justify-between w-full py-1.5 px-2"):
                    ui.label("Saved Bookmarks").classes("text-xs font-bold text-primary")
                    ui.badge(str(len(saved)), color="blue-2", text_color="primary").classes("text-[10px] px-2 py-0.5 rounded-full")
            
            with exp:
                with ui.column().classes("w-full pl-2 gap-1 py-1"):
                    for b_cn in saved:
                        with ui.row().classes("w-full items-center justify-between py-1.5 px-2 hover:bg-gray-100 rounded cursor-pointer transition-colors duration-150").on("click", lambda c=b_cn: open_case(c)):
                            ui.label(b_cn).classes("text-[11px] font-medium text-gray-800")

    # ----------------------------- PDF ------------------------------------ #
    def render_pdf():
        pdf_pane.clear()
        with pdf_pane:
            if not state["case"]:
                ui.label("Open a case from the Library  →").classes("text-gray-400 mt-10 w-full text-center")
                return
            cn, fn = state["case"]
            ui.label(cn).classes("text-sm font-bold case-title mb-1")
            src = f"/pdf/{fn}" + (f"#page={state['page']}" if state["page"] else "")
            ui.element("iframe").props(f'src="{src}"').classes("w-full").style(
                "height:calc(100vh - 120px);border:1px solid #e5e7eb;border-radius:6px")

    # --------------------------- Breakdown -------------------------------- #
    def sec(t):
        ui.label(t).classes("sec")

    def chips(items, color, text_color="white"):
        with ui.row().classes("flex-wrap gap-x-2.5 gap-y-3.5 mt-2 mb-3"):
            for it in items:
                ui.chip(it, on_click=lambda term=it: goto(term)).props(f"color={color} text-color={text_color} clickable").classes("cursor-pointer text-sm font-semibold px-3.5 py-1.5 rounded-full")

    def show_graph(cn):
        nodes = [{"id": cn, "label": cn, "color": "#1a73e8", "font": {"bold": True}, "size": 24}]
        edges = []
        
        bd = get_breakdown(cn)
        if bd and bd.get("precedent_index"):
            for p in bd["precedent_index"]:
                cited = p.get("cited_case")
                if cited and cited != NOT_AVAILABLE:
                    if not any(n["id"] == cited for n in nodes):
                        nodes.append({"id": cited, "label": cited, "color": "#00796b", "size": 16})
                    edges.append({"from": cn, "to": cited, "arrows": "to", "label": p.get("treatment", "")})
                    
        children = get_citing_cases(cn)
        for child in children:
            if not any(n["id"] == child for n in nodes):
                nodes.append({"id": child, "label": child, "color": "#d93025", "size": 16})
            edges.append({"from": child, "to": cn, "arrows": "to"})
            
        html_content = f"""
        <script type="text/javascript" src="https://cdnjs.cloudflare.com/ajax/libs/vis-network/9.1.2/standalone/umd/vis-network.min.js"></script>
        <div id="vis_graph" style="width:100%; height:440px; border:1px solid #e5e7eb; border-radius:8px; background-color:#f9fafb;"></div>
        <script>
            var nodes = new vis.DataSet({json.dumps(nodes)});
            var edges = new vis.DataSet({json.dumps(edges)});
            var container = document.getElementById('vis_graph');
            var data = {{ nodes: nodes, edges: edges }};
            var options = {{
                nodes: {{
                    shape: 'dot',
                    font: {{ size: 12, face: 'Plus Jakarta Sans', color: '#111827' }},
                    borderWidth: 2
                }},
                edges: {{
                    width: 1.5,
                    color: {{ color: '#9ca3af', highlight: '#1a73e8' }},
                    font: {{ size: 9, align: 'top', color: '#4b5563' }}
                }},
                physics: {{
                    stabilization: true,
                    barnesHut: {{
                        gravitationalConstant: -1500,
                        centralGravity: 0.3,
                        springLength: 95
                    }}
                }}
            }};
            var network = new vis.Network(container, data, options);
            network.on("doubleClick", function(params) {{
                if (params.nodes.length > 0) {{
                    var clickedNode = params.nodes[0];
                    var el = document.querySelector('.citation-graph-input input');
                    if (el) {{
                        el.value = clickedNode;
                        el.dispatchEvent(new Event('input'));
                    }}
                }}
            }});
        </script>
        """
        
        with ui.dialog() as dialog, ui.card().classes("w-[80vw] max-w-[800px] h-[540px] p-4"):
            with ui.row().classes("w-full items-center justify-between no-wrap mb-2"):
                ui.label(f"Precedent Map — {cn}").classes("text-sm font-bold text-gray-800")
                ui.button(icon="close", on_click=dialog.close).props("flat round dense color=grey-7")
            
            def on_node_select(e):
                if e.value:
                    open_case(e.value)
                    dialog.close()
            graph_input = ui.input(on_change=on_node_select).classes("citation-graph-input hidden")
            
            ui.html(html_content).classes("w-full h-[450px]")
            
        dialog.open()

    async def gen_breakdown(cn):
        breakdown_pane.clear()
        with breakdown_pane:
            ui.label("Breakdown").classes("pane-head")
            with ui.row().classes("items-center gap-2 mt-3"):
                ui.spinner(size="lg")
                ui.label("Analysing with the local model… (~1–2 min)").classes("text-sm")
        if state.get("bd_pending") == cn:
            return  # already generating this case
        state["bd_pending"] = cn
        state.pop("bd_error", None)
        # Generate in a worker thread. analyze_case caches to the DB — and the cache
        # write completes even if THIS client's websocket drops during the wait (the
        # thread keeps running). _finish_breakdown renders now if we're still
        # connected; the _poll_breakdown timer is the backstop that renders from
        # cache after a drop, so the spinner can't get stuck.
        from src.analyze import analyze_case
        state.pop("bd_fresh", None)
        try:
            ca = await run.io_bound(analyze_case, cn, True)  # force: a Regenerate must re-run
            q = ca.quality()
            # Hollow results aren't cached (so the next open re-attempts), so stash
            # this run in session state — render_breakdown shows it with a warning.
            state["bd_fresh"] = (cn, ca.model_dump(mode="json"), q)
        except Exception as e:  # noqa: BLE001
            print(f"Breakdown failed for {cn}: {e}")
            state["bd_error"] = (cn, str(e))
        _finish_breakdown(cn)

    def _finish_breakdown(cn):
        if state.get("bd_pending") != cn:
            return  # already finished — avoid a double render from the timer
        state["bd_pending"] = None
        err = state.get("bd_error")
        if err and err[0] == cn:
            state.pop("bd_error", None)
            try:
                ui.notify(f"Breakdown failed: {err[1]}", type="negative")
            except Exception:
                pass
        if state.get("case") and state["case"][0] == cn:
            try:
                render_breakdown()
            except Exception as e:  # noqa: BLE001
                print(f"[gen_breakdown] render skipped: {e}")

    def _poll_breakdown():
        """Per-client 2 s timer: if the pending breakdown has finished (cached) or
        failed, render it — covers the case where this client dropped mid-wait so the
        awaited render above never ran."""
        cn = state.get("bd_pending")
        if not cn:
            return
        if (state.get("bd_error") or (None,))[0] == cn or get_breakdown(cn) \
                or (state.get("bd_fresh") or (None,))[0] == cn:
            _finish_breakdown(cn)

    def call_chatbot_api(cn, query):
        from src.config import REPO_ROOT, settings
        from src.ingest import extract_pages
        from src.analyze import _chat, _fit_to_context
        import sqlite3

        # Check memory cache first
        case_texts = state.setdefault("case_texts", {})
        if cn not in case_texts:
            con = sqlite3.connect(settings.sqlite_path)
            row = con.execute("SELECT filename FROM judgements WHERE case_no=? LIMIT 1", (cn,)).fetchone()
            con.close()
            if not row:
                raise RuntimeError(f"Judgment file not found for case: {cn}")
                
            pdf_path = REPO_ROOT / "data" / "sc_judgements" / row[0]
            pages = extract_pages(str(pdf_path), ocr_langs=settings.tesseract_langs)
            text = "\n".join(pages)
            fitted = _fit_to_context(text)
            case_texts[cn] = fitted
            
        judgment_text = case_texts[cn]
        
        system_text = (
            "You are a helpful, professional legal research assistant.\n"
            "You are provided with the text of a Supreme Court judgment below.\n"
            "Answer the user's questions about this judgment accurately, objectively, and based strictly on the judgment text.\n"
            "If the answer cannot be found or inferred from the text, state that you do not have enough information.\n\n"
            "Judgment Text:\n"
            f"{judgment_text}"
        )
        
        messages = state["chats"].get(cn, [])
        history_str = ""
        for msg in messages[:-1]: # exclude the latest query
            role = "User" if msg["role"] == "user" else "Assistant"
            history_str += f"{role}: {msg['content']}\n"
            
        user_text = ""
        if history_str:
            user_text += f"Conversation history so far:\n{history_str}\n"
        user_text += f"Latest User Question: {query}"
        
        reply = _chat(system_text, user_text)
        return reply

    @ui.refreshable
    def floating_chat_widget():
        if not state["case"]:
            return
            
        cn = state["case"][0]
        
        # Floating Toggle Button (elongated, bottom-right)
        btn_text = "Close" if state["chat_open"] else "Ask ROS"
        icon = "close" if state["chat_open"] else "chat"
        color = "red" if state["chat_open"] else "primary"
        
        ui.button(btn_text, icon=icon, on_click=lambda: (state.update({"chat_open": not state["chat_open"]}), floating_chat_widget.refresh())).props(
            f"color={color} rounded"
        ).style(
            "position: fixed; bottom: 50px; right: 24px; z-index: 9999; padding: 0 20px; height: 46px; font-size: 13px; font-weight: bold; text-transform: uppercase; letter-spacing: 0.05em; border-radius: 23px; box-shadow: 0 4px 16px rgba(0,0,0,0.2);"
        )
        
        # Floating Chat Dialog Card (opens above the button)
        if state["chat_open"]:
            with ui.card().style(
                "position: fixed; bottom: 120px; right: 24px; width: 360px; height: 480px; z-index: 9999; "
                "border-radius: 16px; border: 1px solid #dadce0; box-shadow: 0 8px 32px rgba(0,0,0,0.15); "
                "background: #ffffff; display: flex; flex-direction: column; overflow: hidden;"
            ).classes("p-0"):
                # Header row
                with ui.row().classes("w-full items-center justify-between bg-primary text-white p-3").style("flex-shrink: 0;"):
                    with ui.row().classes("items-center gap-2"):
                        ui.icon("chat_bubble_outline", size="20px")
                        with ui.column().classes("gap-1"):
                            ui.label("Ask ROS about case:").classes("text-[10px] opacity-90 uppercase tracking-wider font-bold leading-none")
                            ui.label(cn).classes("text-xs font-bold leading-none text-amber-300 truncate").style("max-width: 220px;")
                    ui.button(icon="close", on_click=lambda: (state.update({"chat_open": False}), floating_chat_widget.refresh())).props("flat round dense color=white").classes("text-xs")
                
                # Chat History Area
                messages = state.setdefault("chats", {}).setdefault(cn, [])
                scroll = ui.scroll_area().classes("flex-grow w-full p-3 bg-gray-50")
                with scroll:
                    if not messages:
                        ui.label("Ask me anything about this judgement. For example:\n- What are the main issues in this case?\n- Who was the appellant and what did they argue?\n- Summarize the final decision.").classes("text-gray-400 text-xs whitespace-pre-line p-2 leading-relaxed")
                    else:
                        for msg in messages:
                            if msg["role"] == "user":
                                ui.chat_message(msg["content"], sent=True, name="You").classes("text-xs")
                            else:
                                ui.chat_message(msg["content"], sent=False, name="ROS AI").classes("text-xs")
                                
                    if state.get("chat_loading") == cn:
                        with ui.row().classes("items-center gap-2 pl-2 mt-2"):
                            ui.spinner(size="sm")
                            ui.label("ROS AI is thinking...").classes("text-xs text-gray-500 italic")
                
                scroll.scroll_to(percent=1.0)
                
                # Bottom Input Area
                with ui.row().classes("w-full items-center gap-2 p-2 bg-white border-t no-wrap").style("border-color: #dadce0; flex-shrink: 0;"):
                    chat_input = ui.input(placeholder="Ask a question about this case...").props("outlined dense").classes("flex-grow text-xs")
                    
                    async def on_send():
                        val = chat_input.value.strip()
                        if not val:
                            return
                        chat_input.value = ""
                        messages.append({"role": "user", "content": val})
                        state["chat_loading"] = cn
                        floating_chat_widget.refresh()
                        
                        try:
                            reply = await run.io_bound(call_chatbot_api, cn, val)
                            messages.append({"role": "assistant", "content": reply})
                        except Exception as e:
                            messages.append({"role": "assistant", "content": f"Error calling AI: {e}"})
                        finally:
                            state["chat_loading"] = None
                            floating_chat_widget.refresh()
                            
                    chat_input.on("keydown.enter", on_send)
                    ui.button(icon="send", on_click=on_send).props("flat round dense color=primary").classes("hover:bg-blue-50")
                    
                    def clear_history():
                        state["chats"][cn] = []
                        floating_chat_widget.refresh()
                    ui.button(icon="delete_outline", on_click=clear_history).props("flat round dense color=grey-7").classes("hover:bg-gray-100")

    def render_breakdown():
        breakdown_pane.clear()
        with breakdown_pane:
            ui.label("Analysis & Breakdown").classes("pane-head")
            if not state["case"]:
                ui.label("Select a case from the library to view its analysis.").classes("text-gray-400 mt-4 text-sm")
                return
            
            cn = state["case"][0]
            m = case_meta(cn)
            
            bench = bench_for(cn, m)
            authoring_judges = m.get("judges") or []
            decided_date = m.get("date") or "Date not available"
            
            bench_str = ", ".join(bench) if bench else "Bench information not available in source"
            authoring_str = ", ".join(authoring_judges) if authoring_judges else "Authoring judge not specified"
            
            username = _safe_username()
            bookmarked = is_bookmarked(username, cn)

            # --- Google Material Header Card ---
            with ui.card().classes("w-full p-4 mb-4").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none; background: #ffffff;"):
                # Bookmark Toggle Row
                with ui.row().classes("w-full items-center justify-between no-wrap mb-2"):
                    ui.label("Case Information").classes("text-[10px] uppercase tracking-wider font-bold text-gray-400")
                    def on_bookmark():
                        status = toggle_bookmark(username, cn)
                        bookmark_btn.props(f"icon={'bookmark' if status else 'bookmark_border'}")
                        ui.notify("Added to Bookmarks" if status else "Removed from Bookmarks", color="primary" if status else "grey-7")
                        refresh_bookmarks_ui()
                    bookmark_btn = ui.button(on_click=on_bookmark).props(f"flat round dense icon={'bookmark' if bookmarked else 'bookmark_border'} color=primary").classes("text-sm")
                
                # Bench (Coram) - FIRST
                with ui.row().classes("items-start gap-2 mt-1 w-full no-wrap"):
                    ui.icon("gavel", size="18px", color="primary").classes("mt-0.5 w-5 text-center flex-shrink-0")
                    with ui.column().classes("gap-0.5"):
                        ui.label("Bench (Coram)").classes("text-[10px] uppercase tracking-wider font-bold text-gray-400")
                        ui.label(bench_str).classes("text-xs font-semibold text-gray-800")
                
                # Authoring Judge - SECOND
                with ui.row().classes("items-start gap-2 mt-2 w-full no-wrap"):
                    ui.icon("edit_note", size="18px", color="secondary").classes("mt-0.5 w-5 text-center flex-shrink-0")
                    with ui.column().classes("gap-0.5"):
                        ui.label("Judgement Delivered By").classes("text-[10px] uppercase tracking-wider font-bold text-gray-400")
                        ui.label(authoring_str).classes("text-xs font-semibold text-gray-800")

                # Decided Date - THIRD
                with ui.row().classes("items-start gap-2 mt-2 w-full no-wrap"):
                    ui.icon("calendar_today", size="18px", color="grey-6").classes("mt-0.5 w-5 text-center flex-shrink-0")
                    with ui.column().classes("gap-0.5"):
                        ui.label("Decided On").classes("text-[10px] uppercase tracking-wider font-bold text-gray-400")
                        ui.label(decided_date).classes("text-xs font-semibold text-gray-800")

                ui.separator().classes("my-3")

                # Parties - THIRD (reduced size)
                if m.get("parties") and m["parties"] != NOT_AVAILABLE:
                    with ui.column().classes("gap-0.5"):
                        ui.label("Parties").classes("text-[10px] uppercase tracking-wider font-bold text-gray-400")
                        ui.label(m["parties"]).classes("text-xs font-medium text-gray-700 leading-normal case-title")

            if m.get("keywords"):
                ui.label("Keywords").classes("text-[10px] uppercase font-bold tracking-wider text-gray-400 mt-2 pl-1")
                chips(m["keywords"][:14], "blue-9")

            # Prefer a just-generated result (incl. an uncached hollow one) for
            # this case; otherwise the cached breakdown from the DB.
            fresh = state.get("bd_fresh")
            quality = None
            if fresh and fresh[0] == cn:
                bd, quality = fresh[1], fresh[2]
            else:
                bd = get_breakdown(cn)
            if quality and quality.get("hollow"):
                with ui.card().classes("w-full p-3 mt-2 mb-1").style("border-radius: 10px; border: 1px solid #f0c36d; background: #fff8e6; box-shadow: none;"):
                    with ui.row().classes("items-center gap-2 no-wrap"):
                        ui.icon("warning_amber", color="orange-9", size="20px")
                        with ui.column().classes("gap-0.5"):
                            ui.label("Low-confidence analysis").classes("text-xs font-bold text-orange-9")
                            ui.label(f"The model returned mostly placeholders ({quality['filled']}/{quality['total']} sections filled) — it likely could not digest this judgment. This result was NOT saved; try Regenerate.").classes("text-[11px] text-gray-600 leading-snug")
                    if not demo:
                        ui.button("Regenerate Analysis", on_click=lambda: gen_breakdown(cn)).props("flat dense color=orange-9 icon=refresh").classes("text-xs mt-1")
            if not bd:
                # Basic metadata fallback
                with ui.card().classes("w-full p-4 mt-2").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;"):
                    if m.get("legislation"):
                        sec("Legislation")
                        for s in m["legislation"][:10]:
                            ui.label(f"• {s}").classes("text-sm body-text")
                    
                    ui.separator().classes("my-3")
                    if demo:
                        ui.label("🔒 Full AI analysis (facts · issues · ratio · precedents) is available after login.").classes("text-sm text-gray-500 mb-2")
                        ui.button("Log in to Unlock", on_click=lambda: ui.navigate.to("/login")).props("color=primary rounded unevaluated")
                    else:
                        ui.button("⚡ Generate AI analysis", on_click=lambda: gen_breakdown(cn)).props("color=primary rounded")
                        ui.label("Extracts facts, ratio, citations, and legislation using the local GGUF model.").classes("text-xs text-gray-400 mt-2")
                return

            if bd.get("topics_discussed"):
                ui.label("Topics Discussed").classes("text-[10px] uppercase font-bold tracking-wider text-gray-400 mt-2 pl-1")
                chips(bd["topics_discussed"][:10], "indigo-9")
            
            # Detailed AI Analysis blocks styled as neat cards
            with ui.card().classes("w-full p-4 mt-3").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;"):
                sec("Facts / Factual Matrix")
                ui.label(bd.get("factual_matrix") or "—").classes("body-text mb-3")
                
                if bd.get("legal_issues"):
                    sec("Legal Issues")
                    for li in bd["legal_issues"][:6]:
                        ui.label(f"• {li.get('question') if isinstance(li, dict) else li}").classes("body-text pl-2 mb-1")
                
                sec("Ratio Decidendi")
                ui.label(bd.get("ratio_decidendi") or "—").classes("body-text mb-3")
                
                if bd.get("deciding_factors"):
                    sec("Deciding Factors")
                    for d in bd["deciding_factors"][:8]:
                        ui.label(f"• {d}").classes("body-text pl-2 mb-1")
                
                if bd.get("precedent_index"):
                    sec("Citations & Distinctions")
                    for p in bd["precedent_index"][:12]:
                        cited, tr = p.get("cited_case", ""), p.get("treatment", "")
                        if not cited:
                            continue
                        target = find_case(cited)  # resolves to a case in the repository, if present
                        with ui.row().classes("items-center gap-2 py-1 no-wrap pl-2"):
                            if tr and tr != NOT_AVAILABLE:
                                ui.badge(tr, color="blue-2" if tr in ("Applied", "Followed") else "amber-2", text_color="grey-9").classes("text-[10px] px-2 py-0.5 rounded")
                            if target:
                                ui.link(cited, "#").classes("text-xs font-semibold no-underline text-primary").on("click", lambda t=target: open_case(t))
                                ui.icon("open_in_new", size="13px").classes("text-primary flex-shrink-0").tooltip("Open this judgement in the library")
                            else:
                                ui.link(cited, _web_search_url(cited), new_tab=True).classes("text-xs font-semibold no-underline text-gray-700 hover:text-primary").tooltip("Not in corpus — search the web")
                                ui.icon("travel_explore", size="13px").classes("text-gray-400 flex-shrink-0")
                
                leg = list(dict.fromkeys((m.get("legislation") or []) + (bd.get("legislation_cited") or [])))
                if leg:
                    sec("Legislation Cited")
                    for s in leg[:12]:
                        with ui.row().classes("items-center gap-1.5 py-0.5 no-wrap pl-2"):
                            ui.icon("article", size="15px").classes("text-gray-400 flex-shrink-0")
                            ui.link(s, _web_search_url(s + " Sri Lanka"), new_tab=True).classes("text-xs font-medium text-primary no-underline hover:underline").tooltip("Search the web for this Act / Article")
                
                sec("Final Order")
                ui.label(bd.get("final_order") or "—").classes("body-text mb-2")
                
                if not demo and bd.get("academic_synthesis") and bd["academic_synthesis"] != NOT_AVAILABLE:
                    sec("Scholar's Note")
                    ui.label(bd["academic_synthesis"]).classes("body-text mb-2")
                if not demo:
                    with ui.row().classes("w-full justify-end items-center"):
                        ui.button("Regenerate Analysis", on_click=lambda: gen_breakdown(cn)).props("flat dense color=primary icon=refresh").classes("text-xs")

    # ----------------------------- navigation ----------------------------- #
    def open_case(case_no, page=None):
        con = _con()
        row = con.execute("SELECT case_no, filename FROM judgements WHERE case_no=? LIMIT 1", (case_no,)).fetchone()
        con.close()
        if not row:
            ui.notify(f"Case not found: {case_no}", type="warning")
            return
        state["case"], state["page"] = row, page
        state["workspace_active"] = True
        update_workspace_visibility()
        render_pdf()
        render_breakdown()
        floating_chat_widget.refresh()
        set_active("bd")  # mobile: jump to the analysis when a case opens

    # ----------------------------- Library -------------------------------- #
    def make_year(y, months_dict):
        total_cases = sum(len(cases) for cases in months_dict.values())
        exp = ui.expansion().classes("w-full mb-1").props("dense header-class='bg-gray-50 text-gray-700 text-xs font-semibold rounded-md' expand-icon-class='text-gray-400'")
        with exp.add_slot('header'):
            with ui.row().classes("items-center justify-between w-full py-1.5 px-2"):
                ui.label(y).classes("text-xs font-bold text-gray-700")
                ui.badge(str(total_cases), color="grey-3", text_color="grey-8").classes("text-[10px] px-2 py-0.5 rounded-full")

        loaded = {"v": False}

        def load():
            if loaded["v"] or not exp.value:
                return
            loaded["v"] = True
            with exp:
                with ui.column().classes("w-full pl-2 gap-1 py-1"):
                    def month_sort_key(m_name):
                        try:
                            return _MONTH_NAMES.index(m_name)
                        except ValueError:
                            return 12

                    sorted_months = sorted(months_dict.keys(), key=month_sort_key)
                    for m_name in sorted_months:
                        cases = months_dict[m_name]
                        if not cases:
                            continue

                        # Sub-expansion for each Month
                        m_exp = ui.expansion().classes("w-full pl-1 mb-0.5").props("dense header-class='text-gray-600 text-[11px] font-medium' expand-icon-class='text-gray-400'")
                        with m_exp.add_slot('header'):
                            with ui.row().classes("items-center justify-between w-full py-1 px-1"):
                                ui.label(m_name).classes("text-[11px] font-semibold text-gray-600")
                                ui.badge(str(len(cases)), color="grey-2", text_color="grey-6").classes("text-[9px] px-1.5 py-0.2 rounded-full")

                        with m_exp:
                            with ui.column().classes("w-full pl-2 gap-1 py-1"):
                                for cn, _date in sorted(cases, key=lambda c: c[1], reverse=True):
                                    with ui.row().classes("w-full items-center justify-between py-1.5 px-2 hover:bg-gray-100 rounded cursor-pointer transition-colors duration-150").on("click", lambda c=cn: open_case(c)):
                                        ui.label(cn).classes("text-[11px] font-medium text-gray-800")
                                        if _date:
                                            ui.label(_date).classes("text-[9px] text-gray-400")

        exp.on_value_change(load)

    def show_tree():
        results.clear()
        with results:
            ui.label("Browse by Year & Month").classes("text-xs font-semibold text-gray-500 mb-2 mt-1")
            by = judgements_by_year_month()
            years = sorted((y for y in by if y != "Undated"), reverse=True)
            if "Undated" in by:
                years.append("Undated")
            for y in years:
                make_year(y, by[y])

    def show_results(hits, label):
        results.clear()
        with results:
            with ui.row().classes("items-center justify-between w-full mb-2"):
                ui.label(label).classes("text-xs font-semibold text-gray-500")
                ui.button("Reset Filters", on_click=reset).props("flat dense color=primary icon=restart_alt").classes("text-xs")
            if not hits:
                ui.label("No judgements match these criteria.").classes("text-sm text-gray-500 mt-2")
                return
            for h in hits:
                cn = h["case_no"]
                dt = h.get("date", "")
                snip = h.get("snippet", "")
                why = h.get("why", "")
                with ui.card().classes("w-full p-3 mb-2 cursor-pointer hover:shadow-md transition-shadow duration-200").on("click", lambda c=cn: open_case(c)).style("border-radius: 8px; border: 1px solid #e8eaed; box-shadow: none;"):
                    with ui.row().classes("justify-between items-center w-full no-wrap"):
                        ui.label(cn).classes("text-xs font-bold text-primary truncate").style("max-width: 70%;")
                        with ui.row().classes("items-center gap-1 no-wrap"):
                            if why == "semantic":
                                ui.badge("✦ AI", color="purple-1", text_color="deep-purple").classes("text-[10px] px-2 py-0.5").style("border-radius: 4px; box-shadow: none;").tooltip("Semantically related — the exact words may not appear")
                            elif why == "broad":
                                ui.badge("partial", color="grey-3", text_color="grey-8").classes("text-[10px] px-2 py-0.5").style("border-radius: 4px; box-shadow: none;").tooltip("Matches some of your search terms")
                            if dt:
                                ui.badge(dt[:4], color="blue-1", text_color="primary").classes("text-[10px] px-2 py-0.5").style("border-radius: 4px; box-shadow: none;")
                    if snip:
                        ui.label(snip).classes("text-[11px] text-gray-600 mt-1 line-clamp-2")

    # ---- combinable filters: Justice · legal area · year · month · search ---- #
    _FILTER_ICON = {"judge": "gavel", "area": "category", "year": "event",
                    "month": "calendar_month", "query": "search"}

    def current_filters() -> dict:
        """Live value of every facet control (None when unset)."""
        return {
            "judge": (judge_sel.value or None),
            "area": (area_sel.value or None),
            "year": (year_sel.value or None),
            "month": (month_sel.value or None),
            "query": ((q.value or "").strip() or None),
        }

    def render_active_filters(active):
        active_filters.clear()
        if not active:
            return
        with active_filters:
            ui.label("Filters:").classes("text-[10px] uppercase font-bold tracking-wider text-gray-400 mr-1 self-center")
            for k, v in active:
                disp = f'"{v}"' if k == "query" else (" · ".join(v) if isinstance(v, list) else str(v))
                with ui.row().classes("items-center gap-0.5 bg-blue-1 rounded-full pl-2 pr-0.5 py-0.5 no-wrap"):
                    ui.icon(_FILTER_ICON.get(k, "filter_alt"), size="13px").classes("text-primary")
                    ui.label(disp).classes("text-[10px] font-semibold text-primary truncate").style("max-width: 110px;")
                    ui.button(icon="close", on_click=lambda key=k: clear_one(key)).props("flat round dense size=xs color=primary")

    def clear_one(key):
        state["_resetting"] = True
        if key == "query":
            q.value = ""
        elif key == "judge":
            judge_sel.value = []          # multi-select clears to an empty list
        else:
            {"area": area_sel, "year": year_sel, "month": month_sel}[key].value = None
        state["_resetting"] = False
        apply_filters()

    def apply_filters(deep: bool = False):
        """Re-run the AND-combined facet query; repaint results + active-filter chips.
        deep=True (Enter / ✦ button) also merges AI semantic matches once the
        embedder is warm — typing stays on the fast FTS path."""
        if state.get("_resetting"):
            return
        f = current_filters()
        active = [(k, v) for k, v in f.items() if v]
        render_active_filters(active)
        if not active:
            show_tree()
            return
        state["workspace_active"] = True
        update_workspace_visibility()
        deep = deep and bool(f.get("query")) and embedder_ready()
        hits = combined_search(**f, semantic=deep, limit=200)
        pretty = " · ".join((f'"{v}"' if k == "query" else (" / ".join(v) if isinstance(v, list) else str(v))) for k, v in active)
        if deep:
            pretty += "  ·  ✦ deep"
        show_results(hits, f"{pretty}  ·  {len(hits)} cases")

    def reset_controls_silently():
        state["_resetting"] = True
        judge_sel.value = []             # multi-select resets to an empty list
        for ctrl in (area_sel, year_sel, month_sel):
            ctrl.value = None
        q.value = ""
        state["_resetting"] = False

    def set_area_quick(area):
        """A quick-filter chip just sets the area facet (combines with the rest)."""
        area_sel.value = area  # fires on_change -> apply_filters

    def goto(term):
        """Jump the Library to a topic / keyword / Act clicked in the breakdown."""
        reset_controls_silently()
        q.value = term or ""
        apply_filters()
        set_active("library")  # mobile: show the related results

    def reset():
        reset_controls_silently()
        render_active_filters([])
        show_tree()

    # ------------------------------ layout -------------------------------- #
    with ui.header().classes("items-center justify-between text-white shadow-none").style("background: #154273; border-bottom: 3px solid #01689b; height: 56px;"):
        # Left side containing AI Active and Logout/Login
        with ui.row().classes("items-center gap-3"):
            if demo:
                ui.badge("DEMO · read-only", color="orange").classes("px-3 py-1 text-xs font-semibold").style("border-radius: 2px; box-shadow: none;")
                ui.button("Log in", on_click=lambda: ui.navigate.to("/login")).props("flat dense color=white").classes("text-xs font-semibold")
            else:
                ui.badge("A.I. Active", color="green").classes("px-3 py-1 text-xs font-semibold").style("border-radius: 2px; box-shadow: none;")
                ui.button(icon="logout", on_click=lambda: (app.storage.user.clear(), ui.navigate.to("/login"))).props("flat round dense color=white").classes("hover:bg-white/10")

        # Middle wordmark — white on the navy government band (rechtspraak.nl feel)
        ui.label("ROS").classes("absolute-center text-3xl font-bold").style("font-family: 'Lora', Georgia, serif; letter-spacing: 0.25em; color: #ffffff;")

        # Right side — Ingestion / Extraction portal (authed tool)
        if not demo:
            ui.button("Extract", icon="document_scanner",
                      on_click=lambda: ui.navigate.to("/extractor")) \
                .props("unelevated dense color=white text-color=primary").classes("text-xs font-semibold")

    # mobile-only tab switcher (hidden on desktop via CSS)
    tab_btns: dict = {}
    mobile_tabs_row = ui.row().classes("mobile-tabs w-full items-stretch gap-0 bg-white border-b shadow-none").style("border-color: #dadce0;")
    with mobile_tabs_row:
        for key, label, icon in [("library", "Library", "menu_book"), ("bd", "Breakdown", "gavel"), ("pdf", "Document", "description")]:
            tab_btns[key] = ui.button(label, icon=icon, on_click=lambda k=key: set_active(k)).props("flat no-caps dense").classes("flex-grow")

    panes_row = ui.row().classes("panes-row w-full no-wrap gap-3 p-3 bg-gray-100")
    with panes_row:
        pdf_pane = ui.column().classes("pane w-2/5 h-full overflow-auto p-4 bg-white border rounded shadow-sm")
        breakdown_pane = ui.column().classes("pane w-2/5 h-full overflow-auto p-4 bg-white border rounded shadow-sm")
        library = ui.column().classes("pane w-1/5 h-full overflow-auto p-4 bg-white border rounded shadow-sm gap-3")
        with library:
            # Welcome header (visible only in welcome state)
            welcome_header = ui.column().classes("w-full items-center gap-2 mb-6")
            with welcome_header:
                ui.label("⚖️ ROScribe").classes("text-3xl font-bold text-primary").style("font-family: 'Lora', Georgia, serif; letter-spacing: 0.1em;")
                ui.label("Supreme Court of Sri Lanka — Legal Research Portal").classes("text-[10px] text-gray-500 text-center uppercase tracking-wider font-bold")
            
            with ui.row().classes("w-full items-center justify-between no-wrap"):
                library_title = ui.label("Library").classes("pane-head")
                home_btn = ui.button(icon="home", on_click=lambda: go_home()).props("flat round dense color=grey-7").classes("hover:bg-gray-100")

            containers["bookmarks"] = ui.column().classes("w-full mb-1")
            
            # Google Search Input style
            q = ui.input(placeholder='Search parties, case no, RDA, "exact phrase"…',
                         on_change=lambda e: apply_filters()).props("rounded outlined dense clearable").classes("w-full")
            with q.add_slot('prepend'):
                ui.icon("search").classes("text-gray-400")
            with q.add_slot('append'):
                ui.button(icon="auto_awesome", on_click=lambda: apply_filters(deep=True)) \
                    .props("flat round dense color=primary") \
                    .tooltip("Deep search — also finds semantically related cases (or press Enter)")
            q.on("keydown.enter", lambda: apply_filters(deep=True))
            
            # Quick filter areas chips (horizontal scroll)
            ui.label("Quick Filters").classes("text-[10px] uppercase font-bold tracking-wider text-gray-400 mt-1")
            with ui.scroll_area().classes("w-full h-8 mb-1"):
                with ui.row().classes("no-wrap gap-1"):
                    popular_areas = ["Fundamental Rights", "Land & Property", "Contract", "Criminal Law & Procedure", "Civil Procedure"]
                    for pa in popular_areas:
                        ui.chip(pa, color="blue-1", on_click=lambda term=pa: set_area_quick(term)).props("outline clickable dense").classes("text-[10px] font-semibold text-primary")
            
            judge_sel = ui.select(distinct_justices(), label="By Justice (select one or more)", with_input=True,
                                  multiple=True, clearable=True,
                                  on_change=lambda e: apply_filters()).props("outlined dense rounded").classes("w-full")
            area_sel = ui.select(legal_areas(), label="By legal area", with_input=True, clearable=True,
                                 on_change=lambda e: apply_filters()).props("outlined dense rounded").classes("w-full")
            with ui.row().classes("w-full no-wrap gap-2"):
                year_sel = ui.select(_available_years(), label="Year", with_input=True, clearable=True,
                                     on_change=lambda e: apply_filters()).props("outlined dense rounded").classes("flex-1 min-w-0")
                month_sel = ui.select(_MONTH_NAMES, label="Month", clearable=True,
                                      on_change=lambda e: apply_filters()).props("outlined dense rounded").classes("flex-1 min-w-0")
            active_filters = ui.row().classes("w-full items-center gap-1 mt-1").style("flex-wrap: wrap;")
            results = ui.column().classes("w-full")

    with ui.footer().classes("items-center justify-center bg-white border-t p-1 shadow-none").style("height: 30px; border-color: #dadce0;"):
        ui.label("built by SMS").classes("text-[10px] text-gray-500 font-semibold uppercase tracking-wider")

    def go_home():
        q.value = ""
        state["case"] = None
        state["workspace_active"] = False
        update_workspace_visibility()
        show_tree()
        floating_chat_widget.refresh()

    def update_workspace_visibility():
        active = state.get("workspace_active", False)
        
        pdf_pane.set_visibility(active)
        breakdown_pane.set_visibility(active)
        welcome_header.set_visibility(not active)
        mobile_tabs_row.set_visibility(active)
        
        if active:
            panes_row.classes(remove="justify-center")
            library.classes(remove="w-full max-w-[800px] mx-auto mt-10 p-6 shadow-md")
            library.classes(add="w-1/5 p-4 shadow-sm")
            library_title.classes(remove="text-2xl text-center mb-4")
            library_title.classes(add="pane-head")
            home_btn.set_visibility(True)
        else:
            panes_row.classes(add="justify-center")
            library.classes(remove="w-1/5 p-4 shadow-sm")
            library.classes(add="w-full max-w-[800px] mx-auto mt-10 p-6 shadow-md")
            library_title.classes(remove="pane-head")
            library_title.classes(add="text-2xl text-center mb-4")
            home_btn.set_visibility(False)

    def set_active(name):
        for k, pane in {"pdf": pdf_pane, "bd": breakdown_pane, "library": library}.items():
            pane.classes(add="active") if k == name else pane.classes(remove="active")
        for k, btn in tab_btns.items():
            btn.props(f"color={'primary' if k == name else 'grey-6'}")

    show_tree()
    refresh_bookmarks_ui()
    render_pdf()
    render_breakdown()
    floating_chat_widget()
    update_workspace_visibility()
    set_active("library")  # default visible pane on mobile (desktop shows all 3)
    ui.timer(2.0, _poll_breakdown)  # render a finished breakdown even if the client dropped mid-wait


@ui.page("/", title="ROS", favicon="data/logos/logo_emblem.png")
def home():
    build_workspace(demo=False)


@ui.page("/demo", title="ROS", favicon="data/logos/logo_emblem.png")
def demo_page():
    build_workspace(demo=True)


@ui.page("/extractor", title="Ingestion Portal", favicon="data/logos/logo_emblem.png")
def extractor_page():
    from metadata_extractor.extractor import run_extraction
    from metadata_extractor.models import JudgmentMetadata
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    import logging
    
    logger = logging.getLogger("metadata_extractor_ui")

    # Check authorization first
    if not app.storage.user.get("authenticated", False):
        return RedirectResponse("/login")

    # Page state
    queue = {}
    table_rows = []
    results = []

    def build_excel_bytes(extracted_records: list[dict]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Registry Overview"
        ws.views.sheetView[0].showGridLines = True
        
        headers = [
            "Case Number", 
            "Date of Judgment", 
            "Appellants / Petitioners", 
            "Respondents", 
            "Judges", 
            "Legislation Cited", 
            "Keywords"
        ]
        ws.append(headers)
        
        font_family = "Segoe UI"
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28
        
        data_font = Font(name=font_family, size=10)
        thin_border = Border(
            left=Side(style='thin', color='DADCE0'),
            right=Side(style='thin', color='DADCE0'),
            top=Side(style='thin', color='DADCE0'),
            bottom=Side(style='thin', color='DADCE0')
        )
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_center = Alignment(horizontal="center", vertical="center")
        
        fill_even = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        for row_idx, record in enumerate(extracted_records, 2):
            meta = record.get("metadata", {})
            parties = meta.get("parties", {})
            
            appellants = ", ".join(parties.get("appellants_petitioners", [])) if isinstance(parties.get("appellants_petitioners"), list) else ""
            respondents = ", ".join(parties.get("respondents", [])) if isinstance(parties.get("respondents"), list) else ""
            judges = ", ".join(meta.get("judges", [])) if isinstance(meta.get("judges"), list) else ""
            legislation = ", ".join(meta.get("legislation_cited", [])) if isinstance(meta.get("legislation_cited"), list) else ""
            keywords = ", ".join(meta.get("keywords", [])) if isinstance(meta.get("keywords"), list) else ""
            
            row_data = [
                meta.get("case_number", ""),
                meta.get("date_of_judgment", ""),
                appellants,
                respondents,
                judges,
                legislation,
                keywords
            ]
            ws.append(row_data)
            ws.row_dimensions[row_idx].height = 24
            
            is_even = (row_idx % 2 == 0)
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                if is_even:
                    cell.fill = fill_even
                
                if col_idx in (1, 2):
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                lines = val.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
            
        bytes_io = io.BytesIO()
        wb.save(bytes_io)
        return bytes_io.getvalue()

    # Dynamic styling helpers
    ui.add_head_html("""
    <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700&family=Lora:ital,wght@0,400;0,500;0,600;0,700;1,400&display=swap" rel="stylesheet">
    <style>
      body {
        background: #f8f9fa;
        font-family: 'Plus Jakarta Sans', sans-serif;
      }
    </style>
    """)

    # Header Bar
    with ui.header().classes("items-center justify-between bg-white text-gray-900 border-b shadow-none").style("border-color: #dadce0; height: 56px;"):
        with ui.row().classes("items-center gap-3"):
            ui.button(icon="arrow_back", on_click=lambda: ui.navigate.to("/")).props("flat round dense color=grey-7").classes("hover:bg-gray-100")
            ui.label("Batch Ingestion Portal").classes("text-sm font-bold text-gray-800")
        ui.label("ROS").classes("absolute-center text-3xl font-bold").style("font-family: 'Lora', Georgia, serif; letter-spacing: 0.25em; color: #0f2d59;")

    # Main layout grid (Left: Settings + Upload, Right: Progress & Table)
    with ui.row().classes("w-full no-wrap gap-4 p-4 items-start"):
        
        # Left Panel (Settings + Upload)
        with ui.column().classes("w-1/3 gap-4"):
            
            # Settings Card
            with ui.card().classes("w-full p-4").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;"):
                ui.label("Extraction Settings").classes("text-xs font-bold text-gray-400 uppercase tracking-wider mb-2")
                
                # Default provider selection based on workspace settings
                env_provider = "llamacpp" if settings.llm_provider.lower() in ("llamacpp", "llamacpp-gguf") else "openai"
                provider_sel = ui.select(
                    {"llamacpp": "Local GGUF (Llama.cpp)", "openai": "OpenAI / Groq API", "anthropic": "Anthropic Claude API"},
                    value=env_provider,
                    label="LLM Provider"
                ).classes("w-full")
                
                # Model or GGUF Path
                env_model = settings.llamacpp_model_path if env_provider == "llamacpp" else settings.llm_model
                model_input = ui.input("Model or GGUF Path", value=env_model).classes("w-full")
                
                # Custom Base URL (Groq/Ollama/OpenRouter)
                env_base_url = settings.openai_base_url if "openai" in settings.llm_provider.lower() else ""
                base_url_input = ui.input("Custom API Base URL (Optional)", value=env_base_url, placeholder="e.g. https://api.groq.com/openai/v1").classes("w-full")
                
                # API Key (hidden by default)
                apikey_input = ui.input("API Key (Optional)", password=True).classes("w-full")
                
                # Concurrency slider
                ui.label("Parallel Processing Threads").classes("text-[10px] text-gray-500 font-semibold mt-2")
                workers_slider = ui.slider(min=1, max=10, value=3).props("label label-always")

                # Auto-change settings helper
                def on_provider_change(e):
                    if e.value == "llamacpp":
                        model_input.value = settings.llamacpp_model_path
                        base_url_input.value = ""
                        apikey_input.value = ""
                        base_url_input.disable()
                        apikey_input.disable()
                    elif e.value == "openai":
                        model_input.value = settings.llm_model if "openai" in settings.llm_provider.lower() else "gpt-4o-mini"
                        base_url_input.value = settings.openai_base_url if "openai" in settings.llm_provider.lower() else ""
                        base_url_input.enable()
                        apikey_input.enable()
                    else:  # anthropic
                        model_input.value = "claude-3-5-haiku-20241022"
                        base_url_input.value = ""
                        base_url_input.disable()
                        apikey_input.enable()
                        
                provider_sel.on_value_change(on_provider_change)
                # Run once initially
                if env_provider == "llamacpp":
                    base_url_input.disable()
                    apikey_input.disable()

            # Upload Card
            with ui.card().classes("w-full p-4").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;"):
                ui.label("Upload Documents").classes("text-xs font-bold text-gray-400 uppercase tracking-wider mb-2")
                
                async def handle_upload(e):
                    name = e.file.name
                    # Prevent duplicate file names in the active queue
                    if any(r["filename"] == name for r in table_rows):
                        ui.notify(f"File already in queue: {name}", color="warning")
                        return
                        
                    content_bytes = await e.file.read()
                    try:
                        if name.lower().endswith(".pdf"):
                            import fitz
                            doc = fitz.open(stream=content_bytes, filetype="pdf")
                            text = ""
                            for i, page in enumerate(doc):
                                text += f"\n===== Page {i+1} =====\n" + page.get_text()
                        else:
                            text = content_bytes.decode("utf-8", errors="ignore")
                        
                        queue[name] = text
                        table_rows.append({"filename": name, "case_number": "—", "date": "—", "status": "Pending"})
                        results_table.update()
                        logger.info(f"File uploaded successfully: {name}")
                        print(f"File uploaded successfully: {name}", flush=True)
                        ui.notify(f"File successfully added to queue: {name}", color="positive")
                    except Exception as err:
                        logger.error(f"Failed to read file {name}: {err}")
                        print(f"Failed to read file {name}: {err}", flush=True)
                        ui.notify(f"Failed to read {name}: {err}", color="negative")
                        
                ui.upload(multiple=True, label="Drag & Drop Files (.txt, .pdf)", auto_upload=True, on_upload=handle_upload).classes("w-full")

        # Right Panel (Queue & Progress Grid)
        with ui.column().classes("w-2/3 gap-4"):
            
            # Actions & Stats Row
            with ui.row().classes("w-full items-center justify-between p-4 bg-white border").style("border-radius: 12px; border-color: #dadce0;"):
                with ui.row().classes("gap-2"):
                    # Process Button
                    async def start_ingestion():
                        if not queue:
                            ui.notify("Queue is empty. Upload some files first.", color="warning")
                            return
                            
                        ingest_btn.props("disable")
                        clear_btn.props("disable")
                        export_btn.props("disable")
                        
                        # Set active state
                        for r in table_rows:
                            if r["status"] == "Pending":
                                r["status"] = "Queued"
                        results_table.update()
                        
                        from nicegui import run
                        import asyncio
                        
                        max_workers = int(workers_slider.value)
                        sem = asyncio.Semaphore(max_workers)
                        
                        async def process_file(fname, txt):
                            row = next(r for r in table_rows if r["filename"] == fname)
                            if not (row["status"] in ("Queued", "Pending", "Failed") or "Failed" in row["status"]):
                                return None  # Skip already completed files
                                
                            async with sem:
                                row["status"] = "Extracting..."
                                results_table.update()
                                
                                try:
                                    meta = await run.io_bound(
                                        run_extraction,
                                        text=txt,
                                        provider=provider_sel.value,
                                        model_or_path=model_input.value,
                                        api_key=apikey_input.value or None,
                                        base_url=base_url_input.value or None
                                    )
                                    row["case_number"] = meta.case_number
                                    row["date"] = meta.date_of_judgment
                                    row["status"] = "Completed"
                                    
                                    # Append to results list
                                    # Remove old record if re-running
                                    results[:] = [r for r in results if r["filepath"] != fname]
                                    results.append({
                                        "filepath": fname,
                                        "metadata": meta.model_dump(mode="json")
                                    })
                                    return True
                                except Exception as exc:
                                    row["status"] = f"Failed: {exc}"
                                    return False
                                finally:
                                    results_table.update()
                        
                        tasks = [process_file(fname, txt) for fname, txt in queue.items()]
                        outcomes = await asyncio.gather(*tasks)
                        
                        valid_outcomes = [o for o in outcomes if o is not None]
                        if not valid_outcomes:
                            ui.notify("No new files to process.", color="info")
                        else:
                            success_count = sum(1 for o in valid_outcomes if o)
                            fail_count = len(valid_outcomes) - success_count
                            ui.notify(f"Completed: {success_count} succeeded, {fail_count} failed.", color="positive" if fail_count == 0 else "warning")
                            
                        ingest_btn.props(remove="disable")
                        clear_btn.props(remove="disable")
                        export_btn.props(remove="disable")
                        
                    ingest_btn = ui.button("Start Ingestion", icon="play_arrow", on_click=start_ingestion).props("color=primary rounded")
                    
                    # Clear Button
                    def clear_queue():
                        queue.clear()
                        table_rows.clear()
                        results.clear()
                        results_table.update()
                        ui.notify("Queue cleared", color="grey-7")
                        
                    clear_btn = ui.button("Clear Queue", icon="clear_all", on_click=clear_queue).props("flat rounded color=grey-7")
                
                # Export to Excel Button
                def download_excel():
                    if not results:
                        ui.notify("No completed extractions to export. Run ingestion first.", color="warning")
                        return
                    try:
                        xlsx_data = build_excel_bytes(results)
                        ui.download(xlsx_data, "metadata_registry.xlsx")
                        ui.notify("Excel spreadsheet generated successfully!", color="positive")
                    except Exception as err:
                        ui.notify(f"Excel generation failed: {err}", color="negative")
                        
                export_btn = ui.button("Export to Excel", icon="grid_on", on_click=download_excel).props("color=green rounded")

            # Table Card
            with ui.card().classes("w-full p-4").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;"):
                ui.label("Extraction Queue").classes("text-xs font-bold text-gray-400 uppercase tracking-wider mb-2")
                
                columns = [
                    {"name": "filename", "label": "File Name", "field": "filename", "align": "left"},
                    {"name": "case_number", "label": "Case Number", "field": "case_number", "align": "center"},
                    {"name": "date", "label": "Date of Judgment", "field": "date", "align": "center"},
                    {"name": "status", "label": "Status", "field": "status", "align": "center"}
                ]
                
                results_table = ui.table(columns=columns, rows=table_rows, row_key="filename").classes("w-full shadow-none border")


# Pre-warm the local model in the background at startup so (a) the first breakdown
# isn't slowed by a cold load and (b) two early requests can't race into the model
# loader at once (the race that could wedge the shared-context lock). Non-fatal.
def _prewarm_model():
    import threading
    def _load():
        if settings.llm_provider.lower() == "llamacpp":
            try:
                from src.analyze import _get_llama
                _get_llama()
                print("[prewarm] local model ready")
            except Exception as e:  # noqa: BLE001
                print(f"[prewarm] model preload skipped: {e}")
        # Then the semantic embedder (bge-m3) so deep search is instant; the two
        # fit together in RAM (3B + bge-m3 — do NOT prewarm a 14B alongside).
        try:
            from src.store import warm_embedder
            if warm_embedder():
                print("[prewarm] semantic embedder ready")
        except Exception as e:  # noqa: BLE001
            print(f"[prewarm] embedder preload skipped: {e}")
    threading.Thread(target=_load, name="model-prewarm", daemon=True).start()


_prewarm_model()

ui.run(host="127.0.0.1", port=8080, show=False, reload=False,
       storage_secret=STORAGE_SECRET, title="ROS", favicon="data/logos/logo_emblem.png")

