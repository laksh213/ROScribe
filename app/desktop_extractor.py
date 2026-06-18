import os
import sys
import shutil
from pathlib import Path

# 1. Determine persistent user data directory
def get_app_data_dir() -> Path:
    if sys.platform == 'win32':
        base = Path(os.environ.get('APPDATA', str(Path.home() / 'AppData' / 'Roaming')))
    elif sys.platform == 'darwin':
        base = Path.home() / 'Library' / 'Application Support'
    else:
        base = Path.home() / '.config'
    app_dir = base / "ROSExtractor"
    app_dir.mkdir(parents=True, exist_ok=True)
    return app_dir

app_data = get_app_data_dir()

# 2. Check if running in a PyInstaller frozen state
if getattr(sys, 'frozen', False):
    bundle_dir = Path(sys._MEIPASS)
    # Honour a user-provided .env beside the app (or in app data) first, so a
    # lean build with no baked-in model can still point at the user's own GGUF
    # path / API keys via the in-UI field or environment.
    try:
        from dotenv import load_dotenv
        for _cand in (Path(sys.executable).resolve().parent / ".env", app_data / ".env"):
            if _cand.exists():
                load_dotenv(_cand)
                break
    except Exception:
        pass
    # Use the bundled model only if one was actually packaged; otherwise leave
    # LLAMACPP_MODEL_PATH to whatever .env/env provides (previously this pointed
    # at a non-existent bundle_dir/model.gguf and broke lean builds).
    model_bundled = bundle_dir / "data" / "model.gguf"
    if model_bundled.exists():
        os.environ["LLAMACPP_MODEL_PATH"] = str(model_bundled)
        
    (app_data / "data").mkdir(parents=True, exist_ok=True)
    os.environ.setdefault("SQLITE_PATH", str(app_data / "data" / "roscribe.db"))
    os.environ["CHROMA_DIR"] = str(bundle_dir / "data" / "chroma")
else:
    # Development mode: resolve path and load .env if available
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    from dotenv import load_dotenv
    load_dotenv()

# Configure sys.path so we can import project modules correctly
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from nicegui import ui
from src.config import settings

# Import extraction logic
from metadata_extractor.extractor import run_extraction
from metadata_extractor.models import JudgmentMetadata
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("desktop_extractor")

@ui.page("/")
def extractor_page():
    # Page state
    queue = {}
    table_rows = []
    results = []
    upload_state = {"count": 0, "timer": None}

    def refresh_table():
        results_table.rows = list(table_rows)
        results_table.update()

    def build_excel_bytes(extracted_records: list[dict]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Registry Overview"
        ws.views.sheetView[0].showGridLines = True
        
        headers = [
            "Case Number", 
            "Date of Judgment", 
            "Appellants / Petitioners", 
            "Respondents", 
            "Judges", 
            "Legislation Cited", 
            "Keywords"
        ]
        ws.append(headers)
        
        font_family = "Segoe UI"
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28
        
        data_font = Font(name=font_family, size=10)
        thin_border = Border(
            left=Side(style='thin', color='DADCE0'),
            right=Side(style='thin', color='DADCE0'),
            top=Side(style='thin', color='DADCE0'),
            bottom=Side(style='thin', color='DADCE0')
        )
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_center = Alignment(horizontal="center", vertical="center")
        
        fill_even = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        for row_idx, record in enumerate(extracted_records, 2):
            meta = record.get("metadata", {})
            parties = record.get("metadata", {}).get("parties", {})
            
            appellants = ", ".join(parties.get("appellants_petitioners", [])) if isinstance(parties.get("appellants_petitioners"), list) else ""
            respondents = ", ".join(parties.get("respondents", [])) if isinstance(parties.get("respondents"), list) else ""
            judges = ", ".join(meta.get("judges", [])) if isinstance(meta.get("judges"), list) else ""
            leg_list = meta.get("legislation_cited", [])
            legislation = ", ".join(str(x) for x in leg_list) if isinstance(leg_list, list) else ""
            keywords = ", ".join(meta.get("keywords", [])) if isinstance(meta.get("keywords"), list) else ""
            
            row_data = [
                meta.get("case_number", ""),
                meta.get("date_of_judgment", ""),
                appellants,
                respondents,
                judges,
                legislation,
                keywords
            ]
            ws.append(row_data)
            ws.row_dimensions[row_idx].height = 24
            
            is_even = (row_idx % 2 == 0)
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                if is_even:
                    cell.fill = fill_even
                
                if col_idx in (1, 2):
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                lines = val.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
            
        bytes_io = io.BytesIO()
        wb.save(bytes_io)
        return bytes_io.getvalue()

    # Dynamic styling helpers
    ui.add_head_html("""
    <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700&family=Lora:ital,wght@0,400;0,500;0,600;0,700;1,400&display=swap" rel="stylesheet">
    <style>
      body {
        background: #f8f9fa;
        font-family: 'Plus Jakarta Sans', sans-serif;
      }
    </style>
    """)

    # Header Bar
    with ui.header().classes("items-center justify-between bg-white text-gray-900 border-b shadow-none").style("border-color: #dadce0; height: 56px;"):
        with ui.row().classes("items-center gap-3"):
            ui.icon("batch_prediction").classes("text-xl text-primary pl-2")
            ui.label("Batch Ingestion Portal").classes("text-sm font-bold text-gray-800")
        ui.label("ROS").classes("absolute-center text-3xl font-bold").style("font-family: 'Lora', Georgia, serif; letter-spacing: 0.25em; color: #0f2d59;")

    # Main layout grid (Left: Settings + Upload, Right: Progress & Table)
    with ui.row().classes("w-full no-wrap gap-4 p-4 items-start"):
        
        # Left Panel (Settings + Upload)
        with ui.column().classes("w-1/3 gap-4"):
            
            # Settings Card
            with ui.card().classes("w-full p-4").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;"):
                ui.label("Extraction Settings").classes("text-xs font-bold text-gray-400 uppercase tracking-wider mb-2")
                
                # Default provider selection based on workspace settings
                env_provider = "llamacpp"
                provider_sel = ui.select(
                    {"llamacpp": "Local GGUF (Llama.cpp)", "openai": "OpenAI / Groq API", "anthropic": "Anthropic Claude API"},
                    value=env_provider,
                    label="LLM Provider"
                ).classes("w-full")
                
                # Model or GGUF Path
                env_model = settings.llamacpp_model_path if env_provider == "llamacpp" else settings.llm_model
                model_input = ui.input("Model or GGUF Path", value=env_model).classes("w-full")
                
                # Custom Base URL (Groq/Ollama/OpenRouter)
                env_base_url = settings.openai_base_url if "openai" in settings.llm_provider.lower() else ""
                base_url_input = ui.input("Custom API Base URL (Optional)", value=env_base_url, placeholder="e.g. https://api.groq.com/openai/v1").classes("w-full")
                
                # API Key (hidden by default)
                apikey_input = ui.input("API Key (Optional)", password=True).classes("w-full")
                
                # Concurrency slider — only meaningful for API providers; a local
                # GGUF shares one model instance and always runs sequentially.
                ui.label("Parallel Requests (API providers)").classes("text-[10px] text-gray-500 font-semibold mt-2")
                workers_slider = ui.slider(min=1, max=10, value=3).props("label label-always")
                workers_hint = ui.label("").classes("text-[10px] text-gray-400")

                def _sync_workers_ui(provider):
                    if provider == "llamacpp":
                        workers_slider.disable()
                        workers_hint.set_text("Local GGUF runs one document at a time.")
                    else:
                        workers_slider.enable()
                        workers_hint.set_text("Processes up to N judgments in parallel.")

                # Auto-change settings helper
                def on_provider_change(e):
                    _sync_workers_ui(e.value)
                    if e.value == "llamacpp":
                        model_input.value = settings.llamacpp_model_path
                        base_url_input.value = ""
                        apikey_input.value = ""
                        base_url_input.disable()
                        apikey_input.disable()
                    elif e.value == "openai":
                        model_input.value = settings.llm_model if "openai" in settings.llm_provider.lower() else "gpt-4o-mini"
                        base_url_input.value = settings.openai_base_url if "openai" in settings.llm_provider.lower() else ""
                        base_url_input.enable()
                        apikey_input.enable()
                    else:  # anthropic
                        model_input.value = "claude-3-5-haiku-20241022"
                        base_url_input.value = ""
                        base_url_input.disable()
                        apikey_input.enable()
                        
                provider_sel.on_value_change(on_provider_change)
                # Run once initially
                _sync_workers_ui(env_provider)
                if env_provider == "llamacpp":
                    base_url_input.disable()
                    apikey_input.disable()

            # Upload Card
            with ui.card().classes("w-full p-4").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;"):
                ui.label("Upload Documents").classes("text-xs font-bold text-gray-400 uppercase tracking-wider mb-2")
                
                async def handle_upload(e):
                    name = e.file.name
                    # Prevent duplicate file names in the active queue
                    if any(r["filename"] == name for r in table_rows):
                        ui.notify(f"File already in queue: {name}", color="warning")
                        return
                        
                    content_bytes = await e.file.read()
                    try:
                        if name.lower().endswith(".pdf"):
                            import fitz
                            doc = fitz.open(stream=content_bytes, filetype="pdf")
                            text = ""
                            for i, page in enumerate(doc):
                                text += f"\n===== Page {i+1} =====\n" + page.get_text()
                        else:
                            text = content_bytes.decode("utf-8", errors="ignore")
                        
                        queue[name] = text
                        table_rows.append({"filename": name, "case_number": "—", "date": "—", "status": "Pending"})
                        refresh_table()
                        logger.info(f"File uploaded successfully: {name}")
                        
                        # Debounced notification logic
                        import asyncio
                        client = ui.context.client
                        upload_state["count"] += 1
                        if upload_state["timer"]:
                            upload_state["timer"].cancel()
                            
                        async def show_summary():
                            await asyncio.sleep(0.5)
                            with client:
                                count = upload_state["count"]
                                if count > 0:
                                    ui.notify(f"Successfully added {count} file{'s' if count > 1 else ''} to the queue.", color="positive")
                                    upload_state["count"] = 0
                            upload_state["timer"] = None
                            
                        upload_state["timer"] = asyncio.create_task(show_summary())
                    except Exception as err:
                        logger.error(f"Failed to read file {name}: {err}")
                        ui.notify(f"Failed to read {name}: {err}", color="negative")
                        
                ui.upload(multiple=True, label="Drag & Drop Folder or Files (.txt, .pdf)", auto_upload=True, on_upload=handle_upload).classes("w-full").props("webkitdirectory directory")

        # Right Panel (Queue & Progress Grid)
        with ui.column().classes("w-2/3 gap-4"):
            
            # Actions & Stats Row
            with ui.row().classes("w-full items-center justify-between p-4 bg-white border").style("border-radius: 12px; border-color: #dadce0;"):
                with ui.row().classes("gap-2"):
                    # Process Button
                    async def start_ingestion():
                        if not queue:
                            ui.notify("Queue is empty. Upload some files first.", color="warning")
                            return
                            
                        # Filter down to tasks that need processing
                        run_files = []
                        for fname, txt in queue.items():
                            row = next(r for r in table_rows if r["filename"] == fname)
                            if row["status"] in ("Queued", "Pending", "Failed") or "Failed" in row["status"]:
                                run_files.append((fname, txt))
                                
                        if not run_files:
                            ui.notify("No new files to process.", color="info")
                            return
                            
                        ingest_btn.props("disable")
                        clear_btn.props("disable")
                        export_btn.props("disable")
                        
                        # Set active state
                        for fname, _ in run_files:
                            row = next(r for r in table_rows if r["filename"] == fname)
                            row["status"] = "Queued"
                        refresh_table()
                        
                        # Show progress container
                        progress_container.visible = True
                        progress_label.set_text(f"Processing 0 of {len(run_files)} files...")
                        progress_bar.set_value(0.0)
                        
                        from nicegui import run
                        import asyncio
                        
                        # Local GGUF shares one model instance under a global lock,
                        # so extra workers would only queue. Parallelism helps API
                        # providers (independent HTTP requests) only.
                        max_workers = 1 if provider_sel.value == "llamacpp" else int(workers_slider.value)
                        sem = asyncio.Semaphore(max_workers)
                        
                        completed_tasks = 0
                        total_tasks = len(run_files)
                        success_count = 0
                        fail_count = 0
                        
                        async def process_file(fname, txt):
                            nonlocal completed_tasks, success_count, fail_count
                            row = next(r for r in table_rows if r["filename"] == fname)
                            
                            async with sem:
                                row["status"] = "Extracting..."
                                refresh_table()
                                
                                try:
                                    meta = await run.io_bound(
                                        run_extraction,
                                        text=txt,
                                        provider=provider_sel.value,
                                        model_or_path=model_input.value,
                                        api_key=apikey_input.value or None,
                                        base_url=base_url_input.value or None
                                    )
                                    row["case_number"] = meta.case_number
                                    row["date"] = meta.date_of_judgment
                                    row["status"] = "Completed"
                                    
                                    # Append to results list
                                    results[:] = [r for r in results if r["filepath"] != fname]
                                    results.append({
                                        "filepath": fname,
                                        "metadata": meta.model_dump(mode="json")
                                    })
                                    success_count += 1
                                    return True
                                except Exception as exc:
                                    row["status"] = f"Failed: {exc}"
                                    fail_count += 1
                                    return False
                                finally:
                                    completed_tasks += 1
                                    progress_label.set_text(f"Processing {completed_tasks} of {total_tasks} files...")
                                    progress_bar.set_value(completed_tasks / total_tasks)
                                    refresh_table()
                         
                        tasks = [process_file(fname, txt) for fname, txt in run_files]
                        await asyncio.gather(*tasks)
                        
                        ui.notify(f"Completed: {success_count} succeeded, {fail_count} failed.", color="positive" if fail_count == 0 else "warning")
                        
                        # Hide progress container after a short delay
                        await asyncio.sleep(2.0)
                        progress_container.visible = False
                        
                        ingest_btn.props(remove="disable")
                        clear_btn.props(remove="disable")
                        export_btn.props(remove="disable")
                        
                    ingest_btn = ui.button("Start Ingestion", icon="play_arrow", on_click=start_ingestion).props("color=primary rounded")
                    
                    # Clear Button
                    def clear_queue():
                        queue.clear()
                        table_rows.clear()
                        results.clear()
                        refresh_table()
                        ui.notify("Queue cleared", color="grey-7")
                        
                    clear_btn = ui.button("Clear Queue", icon="clear_all", on_click=clear_queue).props("flat rounded color=grey-7")
                
                # Export to Excel Button
                def download_excel():
                    if not results:
                        ui.notify("No completed extractions to export. Run ingestion first.", color="warning")
                        return
                    try:
                        xlsx_data = build_excel_bytes(results)
                        ui.download(xlsx_data, "metadata_registry.xlsx")
                        ui.notify("Excel spreadsheet generated successfully!", color="positive")
                    except Exception as err:
                        ui.notify(f"Excel generation failed: {err}", color="negative")
                        
                export_btn = ui.button("Export to Excel", icon="grid_on", on_click=download_excel).props("color=green rounded")

            # Progress Container Card
            with ui.card().classes("w-full p-4").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;") as progress_container:
                progress_container.visible = False
                progress_label = ui.label("Ingesting documents...").classes("text-xs font-semibold text-gray-600 mb-2")
                progress_bar = ui.linear_progress(value=0.0).props("stripe lstrip").classes("w-full")

            # Table Card
            with ui.card().classes("w-full p-4").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;"):
                ui.label("Extraction Queue").classes("text-xs font-bold text-gray-400 uppercase tracking-wider mb-2")
                
                columns = [
                    {"name": "filename", "label": "File Name", "field": "filename", "align": "left"},
                    {"name": "case_number", "label": "Case Number", "field": "case_number", "align": "center"},
                    {"name": "date", "label": "Date of Judgment", "field": "date", "align": "center"},
                    {"name": "status", "label": "Status", "field": "status", "align": "center"}
                ]
                
                results_table = ui.table(columns=columns, rows=table_rows, row_key="filename").classes("w-full shadow-none border")

# Pre-warm local GGUF model if applicable
def _prewarm_model():
    import threading
    def _load():
        if settings.llm_provider.lower() == "llamacpp":
            try:
                from src.analyze import _get_llama
                _get_llama()
                print("[prewarm] local model ready", flush=True)
            except Exception as e:
                print(f"[prewarm] model preload skipped: {e}", flush=True)
    threading.Thread(target=_load, name="model-prewarm", daemon=True).start()

_prewarm_model()

# Start NiceGUI in native desktop mode
ui.run(
    reload=False,
    native=True,
    window_size=(1200, 800),
    title="ROS Metadata Extractor",
    favicon="data/logos/logo_emblem.png"
)
