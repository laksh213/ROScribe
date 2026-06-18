"""ROS Metadata Extractor V2 — desktop registry application.

New in V2 (vs app/desktop_extractor.py):
  * 7-column registry format shown in BOTH the UI table and the Excel export:
      Date | Case No | Parties | Judgement By | Keywords | Legislation | Document Attached
  * Per-row "Document Attached" with Open-file and Reveal-in-folder actions
    (desktop-native); the source file is linked in the Excel export too.
  * Username captured at startup and logged into the registry/Excel banner.
  * "Updates" menu: open the GitHub repo or run `git pull` from a source checkout.
  * Reuses the tested extraction engine + provider-aware concurrency from V1.
"""
import os
import sys
import io
import subprocess
import logging
import webbrowser
from pathlib import Path
from datetime import datetime

# ----------------------------------------------------------------------------
# Paths / environment (mirror V1's frozen-vs-dev handling)
# ----------------------------------------------------------------------------
def get_app_data_dir() -> Path:
    if sys.platform == "win32":
        base = Path(os.environ.get("APPDATA", str(Path.home() / "AppData" / "Roaming")))
    elif sys.platform == "darwin":
        base = Path.home() / "Library" / "Application Support"
    else:
        base = Path.home() / ".config"
    d = base / "ROSExtractorV2"
    d.mkdir(parents=True, exist_ok=True)
    return d

APP_DATA = get_app_data_dir()
ATTACHMENTS_DIR = APP_DATA / "attachments"
ATTACHMENTS_DIR.mkdir(parents=True, exist_ok=True)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
REPO_URL = "https://github.com/laksh213/ROScribe"

if getattr(sys, "frozen", False):
    bundle_dir = Path(sys._MEIPASS)  # type: ignore[attr-defined]
    DATA_DIR = bundle_dir / "data"
    try:
        from dotenv import load_dotenv
        for _cand in (Path(sys.executable).resolve().parent / ".env", APP_DATA / ".env"):
            if _cand.exists():
                load_dotenv(_cand)
                break
    except Exception:
        pass
    # PyInstaller's --add-data "<file>:data/model.gguf" places the model *inside*
    # a model.gguf/ directory under its original name, so accept either layout.
    model_bundled = DATA_DIR / "model.gguf"
    if model_bundled.is_file():
        os.environ["LLAMACPP_MODEL_PATH"] = str(model_bundled)
    elif model_bundled.is_dir():
        _gg = sorted(model_bundled.glob("*"))
        if _gg:
            os.environ["LLAMACPP_MODEL_PATH"] = str(_gg[0])
    if os.environ.get("LLAMACPP_MODEL_PATH"):
        # a self-contained build ships its own model — default the UI to Local GGUF
        os.environ.setdefault("LLM_PROVIDER", "llamacpp")
    os.environ["CHROMA_DIR"] = str(DATA_DIR / "chroma")
else:
    sys.path.insert(0, str(PROJECT_ROOT))
    DATA_DIR = PROJECT_ROOT / "data"
    from dotenv import load_dotenv
    load_dotenv()

sys.path.insert(0, str(PROJECT_ROOT))

from nicegui import ui, run
from nicegui import app as ng_app
from src.config import settings
from metadata_extractor.extractor import run_extraction

# serve the logo assets so they can be shown in the UI header (works in both dev
# and frozen builds — DATA_DIR points at the bundle's data/ when packaged)
ng_app.add_static_files("/logos", str(DATA_DIR / "logos"))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("extractor_v2")

REGISTRY_HEADERS = ["Date", "Case No", "Parties", "Judgement By", "Keywords", "Legislation", "Document Attached"]


# ----------------------------------------------------------------------------
# Pure helpers (module-level so they are unit-testable without the GUI)
# ----------------------------------------------------------------------------
def format_parties(parties: dict) -> str:
    """Render the parties dict as 'Appellants v. Respondents'."""
    if not isinstance(parties, dict):
        return ""
    appellants = ", ".join(parties.get("appellants_petitioners", []) or [])
    respondents = ", ".join(parties.get("respondents", []) or [])
    if appellants and respondents:
        return f"{appellants} v. {respondents}"
    return appellants or respondents


import re

# Canonical SC/CA case-type codes (from the registry format sheet) -> aliases
# that may appear in a raw/extracted case number. Listed most-specific first.
CASE_TYPE_CODES: list[tuple[str, list[str]]] = [
    ("MCR", ["MC REVISION", "MCR"]),
    ("CPA", ["CA/PHC/APN", "PHC APN", "APN", "CPA"]),
    ("PHC", ["PROVINCIAL HIGH COURT", "PHC"]),
    ("LTA", ["LEAVE TO APPEAL", "LTA"]),
    ("HCC", ["HIGH COURT CRIMINAL", "HCC"]),
    ("HCA", ["HABEAS CORPUS", "HABEAS", "HCA"]),
    ("RII", ["RESTITUTIO IN INTEGRUM", "RESTITUTIO", "RII"]),
    ("COC", ["CONTEMPT OF COURT", "CONTEMPT", "COC"]),
    ("BOR", ["LAND/ACQ", "LAND ACQUISITION", "ACQUISITION", "LAND", "ACQ", "BOR"]),
    ("WKF", ["WAKFS", "WAKF", "WAQF", "WKF"]),
    ("INJ", ["INJUNCTION", "INJ"]),
    ("MIS", ["MISCELLANEOUS", "MISC", "MIS"]),
    ("BAL", ["BAIL", "BAL"]),
    ("TRF", ["TRANSFER", "TRF"]),
    ("REV", ["REVISIONS", "REVISION", "REV"]),
    ("ELP", ["ELECTION PETITIONS", "ELECTION PETITION", "ELECTION", "ELP"]),
    ("RTI", ["RIGHT TO INFORMATION", "RTI"]),
    ("REM", ["ADMIRALTY APPEALS", "ADMIRALTY", "REM"]),
    ("EXT", ["EXTRADITION APPEALS", "EXTRADITION", "EXT"]),
    ("EXP", ["EXPULSION", "EXP"]),
    ("TAX", ["TAX"]),
    ("WRT", ["WRIT", "WRT"]),
]
# (alias, code) flattened, longest alias first so specific titles win the match
_CASE_ALIASES = sorted(
    [(re.sub(r"[^A-Z0-9]+", " ", a).strip(), c) for c, al in CASE_TYPE_CODES for a in al],
    key=lambda x: len(x[0]), reverse=True,
)


def _expand_year(y: str) -> str:
    """2-digit year -> 4-digit by prefixing '20' (e.g. '26' -> '2026', '07' -> '2007')."""
    return f"20{y}" if len(y) == 2 else y


def normalize_case_number(raw: str) -> str:
    """Correct an extracted case number to the canonical CA/CODE/NNNN/YYYY format
    from the registry sheet: a 'CA/' prefix, the number zero-padded to 4 digits,
    and a 4-digit year (2-digit -> 20xx). E.g. 'CA/TAX/017/26' -> 'CA/TAX/0017/2026',
    'Writ Application No. 11 of 2026' -> 'CA/WRT/0011/2026'. Case types not on the
    sheet (e.g. an 'SC Appeal' number) are returned unchanged."""
    if not raw or not raw.strip():
        return raw or ""
    if "not available" in raw.lower():
        return raw
    s = re.sub(r"[^A-Z0-9]+", " ", raw.upper())
    code, tail = None, s
    for alias, c in _CASE_ALIASES:
        m = re.search(r"\b" + re.escape(alias) + r"\b", s)
        if m:
            code, tail = c, s[m.end():]
            break
    nums = re.findall(r"\d+", tail) or re.findall(r"\d+", s)
    if code and len(nums) >= 2:
        return f"CA/{code}/{int(nums[0]):04d}/{_expand_year(nums[1])}"
    if code and len(nums) == 1:
        return f"CA/{code}/{int(nums[0]):04d}"
    return raw.strip()


def meta_to_row(meta: dict) -> dict:
    """Map an extracted JudgmentMetadata dict to the 7-column registry row."""
    judges = meta.get("judges", []) or []
    authoring = (meta.get("authoring_judge") or "").strip()
    return {
        "date": meta.get("date_of_judgment", "") or "",
        "case_no": normalize_case_number(meta.get("case_number", "") or ""),
        "parties": format_parties(meta.get("parties", {})),
        "judgement_by": authoring or (judges[0] if judges else ""),
        "keywords": ", ".join(meta.get("keywords", []) or []),
        "legislation": ", ".join(str(x) for x in (meta.get("legislation_cited", []) or [])),
    }


def open_path(path: str, reveal: bool = False) -> None:
    """Open a file with the OS default app, or reveal it in the file manager."""
    p = os.path.abspath(path)
    if not os.path.exists(p):
        raise FileNotFoundError(p)
    if sys.platform == "darwin":
        subprocess.run(["open", "-R", p] if reveal else ["open", p], check=False)
    elif sys.platform == "win32":
        if reveal:
            subprocess.run(["explorer", f"/select,{p}"], check=False)
        else:
            os.startfile(p)  # type: ignore[attr-defined]
    else:
        subprocess.run(["xdg-open", os.path.dirname(p) if reveal else p], check=False)


def check_for_update() -> tuple[bool, str]:
    """Run `git pull --ff-only` from a source checkout. Returns (ok, message)."""
    if getattr(sys, "frozen", False) or not (PROJECT_ROOT / ".git").exists():
        return False, f"This is a packaged build. Download the latest version from {REPO_URL}."
    try:
        out = subprocess.run(
            ["git", "-C", str(PROJECT_ROOT), "pull", "--ff-only"],
            capture_output=True, text=True, timeout=120,
        )
        msg = (out.stdout + "\n" + out.stderr).strip()
        return out.returncode == 0, msg or "Up to date."
    except Exception as e:
        return False, f"Update failed: {e}"


def build_excel_bytes(records: list[dict], username: str) -> bytes:
    """Build the V2 registry workbook (7 columns + username/timestamp banner)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Registry"
    font_family = "Segoe UI"

    # --- banner (logs the user + generation time at the start of the table) ---
    ws.append(["ROS Metadata Registry"])
    ws.cell(row=1, column=1).font = Font(name=font_family, size=14, bold=True, color="0F2D59")
    ws.append([f"Compiled by: {username or 'Unknown'}"])
    ws.append([f"Generated: {datetime.now():%Y-%m-%d %H:%M}"])
    ws.append([])
    for r in (2, 3):
        ws.cell(row=r, column=1).font = Font(name=font_family, size=10, italic=True, color="5F6368")

    header_row = ws.max_row + 1
    ws.append(REGISTRY_HEADERS)

    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    for col_idx in range(1, len(REGISTRY_HEADERS) + 1):
        c = ws.cell(row=header_row, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 28

    data_font = Font(name=font_family, size=10)
    link_font = Font(name=font_family, size=10, color="1A73E8", underline="single")
    thin = Side(style="thin", color="DADCE0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_even = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    for i, rec in enumerate(records):
        row = meta_to_row(rec.get("metadata", {}))
        doc_name = rec.get("doc_name", "") or ""
        values = [
            row["date"], row["case_no"], row["parties"],
            row["judgement_by"], row["keywords"], row["legislation"], doc_name,
        ]
        ws.append(values)
        excel_row = ws.max_row
        is_even = (excel_row % 2 == 0)
        for col_idx in range(1, len(values) + 1):
            c = ws.cell(row=excel_row, column=col_idx)
            c.font = data_font
            c.border = border
            if is_even:
                c.fill = fill_even
            c.alignment = Alignment(
                horizontal="center" if col_idx in (1, 2) else "left",
                vertical="center", wrap_text=True,
            )
        # link the "Document Attached" cell to the source file
        doc_path = rec.get("doc_path")
        if doc_path and os.path.exists(doc_path):
            c = ws.cell(row=excel_row, column=len(values))
            c.hyperlink = Path(doc_path).resolve().as_uri()
            c.font = link_font

    # auto width
    for col in ws.iter_cols(min_row=header_row):
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            for line in str(cell.value or "").split("\n"):
                max_len = max(max_len, len(line))
        ws.column_dimensions[letter].width = min(max(max_len + 3, 12), 45)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def append_to_excel(records: list[dict], username: str, filepath: str) -> int:
    """Append a timestamped batch of records to the registry workbook at filepath,
    creating it (with a styled header row) if it does not exist. Existing rows are
    preserved — each call adds a new '--- Batch added ... ---' separator + rows."""
    from openpyxl import load_workbook
    font_family = "Segoe UI"
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Registry"
        ws.append(REGISTRY_HEADERS)
        hfont = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        hfill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        for ci in range(1, len(REGISTRY_HEADERS) + 1):
            c = ws.cell(row=1, column=ci)
            c.font = hfont
            c.fill = hfill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([f"--- Batch added {ts}  by {username or 'Unknown'} ---"] + [""] * (len(REGISTRY_HEADERS) - 1))
    sep = ws.max_row
    ws.merge_cells(start_row=sep, start_column=1, end_row=sep, end_column=len(REGISTRY_HEADERS))
    sc = ws.cell(row=sep, column=1)
    sc.font = Font(name=font_family, size=10, italic=True, bold=True, color="5F6368")
    sc.fill = PatternFill(start_color="E8EAED", end_color="E8EAED", fill_type="solid")
    sc.alignment = Alignment(horizontal="center", vertical="center")

    data_font = Font(name=font_family, size=10)
    link_font = Font(name=font_family, size=10, color="1A73E8", underline="single")
    thin = Side(style="thin", color="DADCE0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for rec in records:
        row = meta_to_row(rec.get("metadata", {}))
        values = [row["date"], row["case_no"], row["parties"], row["judgement_by"],
                  row["keywords"], row["legislation"], rec.get("doc_name", "") or ""]
        ws.append(values)
        er = ws.max_row
        for ci in range(1, len(values) + 1):
            c = ws.cell(row=er, column=ci)
            c.font = data_font
            c.border = border
            c.alignment = Alignment(horizontal="center" if ci in (1, 2) else "left",
                                    vertical="center", wrap_text=True)
        dp = rec.get("doc_path")
        if dp and os.path.exists(dp):
            c = ws.cell(row=er, column=len(values))
            c.hyperlink = Path(dp).resolve().as_uri()
            c.font = link_font

    for col in ws.iter_cols():
        ml = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ml + 3, 12), 45)

    wb.save(filepath)
    return len(records)


# ----------------------------------------------------------------------------
# UI
# ----------------------------------------------------------------------------
@ui.page("/")
def page():
    state = {"username": ""}
    queue: dict[str, dict] = {}          # filename -> {text, doc_path, doc_name}
    table_rows: list[dict] = []
    results: list[dict] = []
    upload_state = {"count": 0, "timer": None}

    def refresh_table():
        results_table.rows = list(table_rows)
        results_table.update()

    def save_attachment(name: str, content: bytes) -> str:
        safe = f"{datetime.now():%Y%m%d_%H%M%S}_{Path(name).name}"
        dest = ATTACHMENTS_DIR / safe
        dest.write_bytes(content)
        return str(dest)

    def empty_row(name: str, doc_path: str) -> dict:
        return {
            "filename": name, "date": "—", "case_no": "—", "parties": "—",
            "judgement_by": "—", "keywords": "—", "legislation": "—",
            "document": name, "doc_path": doc_path, "status": "Pending",
        }

    # ---- styling ----
    ui.add_head_html("""
    <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700&family=Lora:ital,wght@0,400;0,600;0,700&display=swap" rel="stylesheet">
    <style>body { background:#f8f9fa; font-family:'Plus Jakarta Sans', sans-serif; }</style>
    """)

    # ---- header ----
    with ui.header().classes("items-center justify-between bg-white text-gray-900 border-b shadow-none").style("border-color:#dadce0; height:56px;"):
        with ui.row().classes("items-center gap-3"):
            ui.icon("batch_prediction").classes("text-xl text-primary pl-2")
            ui.label("ROS Extractor System").classes("text-sm font-bold text-gray-800")
        ui.label("ROS").classes("absolute-center text-3xl font-bold").style("font-family:'Lora', Georgia, serif; letter-spacing:0.25em; color:#0f2d59;")
        with ui.row().classes("items-center gap-2 pr-2"):
            user_label = ui.label("User: —").classes("text-xs font-semibold text-gray-600")
            with ui.button(icon="system_update_alt").props("flat round dense color=grey-7"):
                with ui.menu() as _menu:
                    ui.menu_item("Open GitHub (downloads)", on_click=lambda: webbrowser.open(REPO_URL))
                    ui.menu_item("Update now (git pull)", on_click=lambda: run_update())

    def run_update():
        ok, msg = check_for_update()
        ui.notify(("Updated. Restart to apply." if "Already up to date" not in msg else "Already up to date.") if ok else msg,
                  color="positive" if ok else "warning", multi_line=True, close_button="OK")

    # ---- main grid ----
    with ui.row().classes("w-full no-wrap gap-4 p-4 items-start"):
        # Left: settings + upload
        with ui.column().classes("w-1/3 gap-4"):
            with ui.card().classes("w-full p-4").style("border-radius:12px; border:1px solid #dadce0; box-shadow:none;"):
                ui.label("Extraction Settings").classes("text-xs font-bold text-gray-400 uppercase tracking-wider mb-2")
                env_provider = "llamacpp"
                provider_sel = ui.select(
                    {
                        "llamacpp": "Local GGUF (Llama.cpp)",
                        "ollama": "Ollama (Local)",
                        "openai": "OpenAI / Groq API",
                        "anthropic": "Anthropic Claude API"
                    },
                    value=env_provider, label="LLM Provider",
                ).classes("w-full")
                env_model = settings.llamacpp_model_path
                model_input = ui.input("Model or GGUF Path", value=env_model).classes("w-full")
                base_url_input = ui.input("Custom API Base URL (Optional)",
                                          value="",
                                          placeholder="e.g. https://api.groq.com/openai/v1").classes("w-full")
                apikey_input = ui.input("API Key (Optional)", password=True).classes("w-full")

                ui.label("Parallel Requests (API providers)").classes("text-[10px] text-gray-500 font-semibold mt-2")
                workers_slider = ui.slider(min=1, max=10, value=3).props("label label-always")
                workers_hint = ui.label("").classes("text-[10px] text-gray-400")

                def sync_workers_ui(provider):
                    if provider == "llamacpp":
                        workers_slider.disable()
                        workers_hint.set_text("Local GGUF runs one document at a time.")
                    elif provider == "ollama":
                        workers_slider.enable()
                        workers_hint.set_text("Processes up to N judgments in parallel via Ollama.")
                    else:
                        workers_slider.enable()
                        workers_hint.set_text("Processes up to N judgments in parallel.")

                def on_provider_change(e):
                    sync_workers_ui(e.value)
                    if e.value == "llamacpp":
                        model_input.value = settings.llamacpp_model_path
                        base_url_input.value = ""; apikey_input.value = ""
                        base_url_input.disable(); apikey_input.disable()
                    elif e.value == "ollama":
                        model_input.value = settings.llm_model
                        base_url_input.value = settings.ollama_base_url
                        apikey_input.value = ""
                        base_url_input.enable(); apikey_input.disable()
                    elif e.value == "openai":
                        model_input.value = settings.llm_model if "openai" in settings.llm_provider.lower() else "gpt-4o-mini"
                        base_url_input.value = settings.openai_base_url if "openai" in settings.llm_provider.lower() else ""
                        base_url_input.enable(); apikey_input.enable()
                    else:
                        model_input.value = "claude-3-5-haiku-20241022"
                        base_url_input.value = ""; base_url_input.disable(); apikey_input.enable()

                provider_sel.on_value_change(on_provider_change)
                sync_workers_ui(env_provider)
                if env_provider == "llamacpp":
                    base_url_input.disable(); apikey_input.disable()

            with ui.card().classes("w-full p-4").style("border-radius:12px; border:1px solid #dadce0; box-shadow:none;"):
                ui.label("Upload Documents").classes("text-xs font-bold text-gray-400 uppercase tracking-wider mb-2")

                async def handle_upload(e):
                    name = e.file.name
                    if any(r["filename"] == name for r in table_rows):
                        ui.notify(f"File already in queue: {name}", color="warning")
                        return
                    try:
                        content_bytes = await e.file.read()
                        doc_path = save_attachment(name, content_bytes)
                        if name.lower().endswith(".pdf"):
                            import fitz
                            doc = fitz.open(stream=content_bytes, filetype="pdf")
                            text = "".join(f"\n===== Page {i+1} =====\n" + p.get_text() for i, p in enumerate(doc))
                        else:
                            text = content_bytes.decode("utf-8", errors="ignore")
                        queue[name] = {"text": text, "doc_path": doc_path, "doc_name": name}
                        table_rows.append(empty_row(name, doc_path))
                        refresh_table()

                        import asyncio
                        client = ui.context.client
                        upload_state["count"] += 1
                        if upload_state["timer"]:
                            upload_state["timer"].cancel()

                        async def show_summary():
                            await asyncio.sleep(0.5)
                            with client:
                                count = upload_state["count"]
                                if count > 0:
                                    ui.notify(f"Added {count} file{'s' if count > 1 else ''} to the queue.", color="positive")
                                    upload_state["count"] = 0
                            upload_state["timer"] = None

                        upload_state["timer"] = asyncio.create_task(show_summary())
                    except Exception as err:
                        logger.error(f"Failed to read {name}: {err}")
                        ui.notify(f"Failed to read {name}: {err}", color="negative")

                ui.upload(multiple=True, label="Drag & Drop Folder or Files (.txt, .pdf)",
                          auto_upload=True, on_upload=handle_upload).classes("w-full").props("webkitdirectory directory")

        # Right: actions + progress + table
        with ui.column().classes("w-2/3 gap-4"):
            with ui.row().classes("w-full items-center justify-between p-4 bg-white border").style("border-radius:12px; border-color:#dadce0;"):
                with ui.row().classes("gap-2"):
                    async def start_ingestion():
                        if not queue:
                            ui.notify("Queue is empty. Upload some files first.", color="warning")
                            return
                        run_files = []
                        for fname, payload in queue.items():
                            row = next(r for r in table_rows if r["filename"] == fname)
                            if row["status"] in ("Queued", "Pending", "Failed") or "Failed" in row["status"]:
                                run_files.append((fname, payload))
                        if not run_files:
                            ui.notify("No new files to process.", color="info")
                            return

                        ingest_btn.props("disable"); clear_btn.props("disable"); export_btn.props("disable")
                        for fname, _ in run_files:
                            next(r for r in table_rows if r["filename"] == fname)["status"] = "Queued"
                        refresh_table()
                        progress_container.visible = True
                        progress_label.set_text(f"Processing 0 of {len(run_files)} files...")
                        progress_bar.set_value(0.0)

                        import asyncio
                        provider_now = provider_sel.value
                        max_workers = 1 if provider_now == "llamacpp" else int(workers_slider.value)
                        sem = asyncio.Semaphore(max_workers)
                        completed = success = fail = 0
                        total = len(run_files)

                        async def process_file(fname, payload):
                            nonlocal completed, success, fail
                            row = next(r for r in table_rows if r["filename"] == fname)
                            async with sem:
                                row["status"] = "Extracting..."; refresh_table()
                                try:
                                    meta = await run.io_bound(
                                        run_extraction, text=payload["text"], provider=provider_now,
                                        model_or_path=model_input.value,
                                        api_key=apikey_input.value or None,
                                        base_url=base_url_input.value or None,
                                    )
                                    md = meta.model_dump(mode="json")
                                    mapped = meta_to_row(md)
                                    row.update(mapped); row["status"] = "Completed"
                                    results[:] = [r for r in results if r["filepath"] != fname]
                                    results.append({
                                        "filepath": fname, "metadata": md,
                                        "doc_path": payload["doc_path"], "doc_name": payload["doc_name"],
                                        "processed_at": datetime.now().isoformat(timespec="seconds"),
                                    })
                                    success += 1
                                except Exception as exc:
                                    row["status"] = f"Failed: {exc}"; fail += 1
                                finally:
                                    completed += 1
                                    progress_label.set_text(f"Processing {completed} of {total} files...")
                                    progress_bar.set_value(completed / total)
                                    refresh_table()

                        await asyncio.gather(*[process_file(f, p) for f, p in run_files])
                        ui.notify(f"Completed: {success} succeeded, {fail} failed.",
                                  color="positive" if fail == 0 else "warning")
                        await asyncio.sleep(2.0)
                        progress_container.visible = False
                        ingest_btn.props(remove="disable"); clear_btn.props(remove="disable"); export_btn.props(remove="disable")

                    ingest_btn = ui.button("Start Ingestion", icon="play_arrow", on_click=start_ingestion).props("color=primary rounded")

                    def clear_queue():
                        queue.clear(); table_rows.clear(); results.clear(); refresh_table()
                        ui.notify("Queue cleared", color="grey-7")

                    clear_btn = ui.button("Clear Queue", icon="clear_all", on_click=clear_queue).props("flat rounded color=grey-7")

                def download_excel():
                    new = [r for r in results if not r.get("exported")]
                    if not new:
                        ui.notify("No new extractions to add since the last export." if results
                                  else "No completed extractions yet. Run ingestion first.",
                                  color="info" if results else "warning")
                        return
                    try:
                        # Append (not overwrite) to a persistent registry on the Desktop;
                        # each export adds a fresh timestamped batch and keeps prior rows.
                        dest_path = os.path.expanduser("~/Desktop/metadata_registry.xlsx")
                        append_to_excel(new, state["username"], dest_path)
                        for r in new:
                            r["exported"] = True
                        with open(dest_path, "rb") as f:
                            ui.download(f.read(), "metadata_registry.xlsx")
                        ui.notify(f"Added {len(new)} record(s) to metadata_registry.xlsx on your Desktop.", color="positive")
                    except Exception as err:
                        ui.notify(f"Excel save failed: {err}", color="negative")

                export_btn = ui.button("Export to Excel", icon="grid_on", on_click=download_excel).props("color=green rounded")

                def open_local_excel():
                    dest_path = os.path.expanduser("~/Desktop/metadata_registry.xlsx")
                    if os.path.exists(dest_path):
                        try:
                            open_path(dest_path)
                            ui.notify(f"Opened Excel file: {os.path.basename(dest_path)}", color="positive")
                        except Exception as err:
                            ui.notify(f"Could not open Excel file: {err}", color="negative")
                    else:
                        ui.notify("No local Excel registry found on Desktop. Export first.", color="warning")

                open_excel_btn = ui.button("Open Excel", icon="open_in_new", on_click=open_local_excel).props("outline color=green rounded")

            with ui.card().classes("w-full p-4").style("border-radius:12px; border:1px solid #dadce0; box-shadow:none;") as progress_container:
                progress_container.visible = False
                progress_label = ui.label("Ingesting documents...").classes("text-xs font-semibold text-gray-600 mb-2")
                progress_bar = ui.linear_progress(value=0.0).props("stripe").classes("w-full")

            with ui.card().classes("w-full p-4").style("border-radius:12px; border:1px solid #dadce0; box-shadow:none;"):
                ui.label("Extraction Registry").classes("text-xs font-bold text-gray-400 uppercase tracking-wider mb-2")
                columns = [
                    {"name": "date", "label": "Date", "field": "date", "align": "center"},
                    {"name": "case_no", "label": "Case No", "field": "case_no", "align": "center"},
                    {"name": "parties", "label": "Parties", "field": "parties", "align": "left"},
                    {"name": "judgement_by", "label": "Judgement By", "field": "judgement_by", "align": "left"},
                    {"name": "keywords", "label": "Keywords", "field": "keywords", "align": "left"},
                    {"name": "legislation", "label": "Legislation", "field": "legislation", "align": "left"},
                    {"name": "document", "label": "Document Attached", "field": "document", "align": "center"},
                    {"name": "status", "label": "Status", "field": "status", "align": "center"},
                ]
                results_table = ui.table(columns=columns, rows=table_rows, row_key="filename").classes("w-full shadow-none border")
                # render the Document Attached cell as Open + Reveal buttons
                results_table.add_slot("body-cell-document", r"""
                    <q-td :props="props" class="text-center">
                        <template v-if="props.row.doc_path">
                            <q-btn dense flat round color="primary" icon="description"
                                   @click="() => $parent.$emit('opendoc', props.row)">
                                <q-tooltip>Open document</q-tooltip>
                            </q-btn>
                            <q-btn dense flat round color="grey-7" icon="folder_open"
                                   @click="() => $parent.$emit('revealdoc', props.row)">
                                <q-tooltip>Show in folder</q-tooltip>
                            </q-btn>
                        </template>
                        <span v-else class="text-grey">—</span>
                    </q-td>
                """)

                def _open(e, reveal):
                    try:
                        open_path(e.args["doc_path"], reveal=reveal)
                    except Exception as err:
                        ui.notify(f"Could not open document: {err}", color="negative")

                results_table.on("opendoc", lambda e: _open(e, False))
                results_table.on("revealdoc", lambda e: _open(e, True))

    # ---- username capture at startup ----
    def load_usernames() -> list[str]:
        import json
        path = APP_DATA / "usernames.json"
        if path.exists():
            try:
                return json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                pass
        sys_user = os.environ.get("USER") or os.environ.get("USERNAME")
        return [sys_user] if sys_user else []

    def save_username(name: str):
        import json
        if not name or name == "Unknown":
            return
        path = APP_DATA / "usernames.json"
        users = load_usernames()
        if name not in users:
            users.append(name)
            try:
                path.write_text(json.dumps(users), encoding="utf-8")
            except Exception:
                pass

    users_list = load_usernames()
    with ui.dialog().props("persistent maximized") as user_dialog, ui.card().classes("items-center justify-center gap-2 w-full h-full").style("background:#f8f9fa; box-shadow:none;"):
        ui.label("ROS").classes("text-5xl font-bold").style("font-family:'Lora', Georgia, serif; letter-spacing:0.25em; color:#0f2d59;")
        ui.label("CASE METADATA EXTRACTION SYSTEM").classes("text-xs font-bold text-gray-400 uppercase tracking-widest mb-4")
        ui.icon("batch_prediction").classes("text-5xl text-primary")
        ui.label("Welcome to ROS Extractor System").classes("text-2xl font-bold text-gray-800 text-center")
        ui.label("An offline-first desktop ingestion utility. Select a returning name or enter a new one to begin this registry session.").classes("text-sm text-gray-500 text-center mb-2").style("max-width:480px;")
        
        selected_user = None
        if users_list:
            selected_user = ui.select({u: u for u in users_list}, value=users_list[0], label="Select Name").classes("w-72")
            
        name_in = ui.input("Or enter a new name", placeholder="Your name").classes("w-72")

        def begin():
            chosen_name = ""
            if name_in.value:
                chosen_name = name_in.value.strip()
            elif selected_user and selected_user.value:
                chosen_name = selected_user.value.strip()
                
            chosen_name = chosen_name or "Unknown"
            state["username"] = chosen_name
            save_username(chosen_name)
            user_label.set_text(f"User: {state['username']}")
            logger.info(f"Session started by user: {state['username']}")
            user_dialog.close()

        name_in.on("keydown.enter", lambda _: begin())
        ui.button("Open Ingestion Portal", icon="launch", on_click=begin).props("color=primary").classes("w-72 py-3 mt-2 text-base font-bold rounded-lg")

    user_dialog.open()


def main():
    use_native = os.getenv("ROSCRIBE_LOCALHOST", "0") != "1"
    kwargs = {
        "reload": False,
        "native": use_native,
        "title": "ROS Extractor System",
        "favicon": str(DATA_DIR / "logos" / "app_icon_1024.png"),
    }
    if use_native:
        kwargs["window_size"] = (1280, 820)
    else:
        kwargs["host"] = "127.0.0.1"
        kwargs["port"] = 8080
    ui.run(**kwargs)


if __name__ in {"__main__", "__mp_main__"}:
    main()
